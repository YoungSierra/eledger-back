import uuid
from datetime import datetime, timezone, date

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.models.bancos import BanBanco, BanCuenta, BanChequera
from app.models.admin import AdmMoneda
from app.models.contabilidad import CntCuenta
from app.schemas.auth import UsuarioActual
from app.schemas.bancos import (
    BancoCreate, BancoUpdate, BancoResponse,
    CuentaBancariaCreate, CuentaBancariaUpdate, CuentaBancariaResponse,
    ChequerapCreate, ChequeraUpdate, ChequeraResponse,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _enrich_cuenta(db: Session, obj: BanCuenta) -> CuentaBancariaResponse:
    banco = db.query(BanBanco).filter(BanBanco.id == obj.banco_id).first()
    moneda = db.query(AdmMoneda).filter(AdmMoneda.id == obj.moneda_id).first() if obj.moneda_id else None
    cuenta = db.query(CntCuenta).filter(CntCuenta.id == obj.cuenta_contable_id).first() if obj.cuenta_contable_id else None
    return CuentaBancariaResponse(
        id=obj.id, banco_id=obj.banco_id,
        banco_nombre=banco.nombre if banco else None,
        nombre=obj.nombre, numero=obj.numero, tipo=obj.tipo,
        moneda_id=obj.moneda_id,
        moneda_codigo=moneda.codigo if moneda else None,
        cuenta_contable_id=obj.cuenta_contable_id,
        cuenta_contable_codigo=cuenta.codigo if cuenta else None,
        cuenta_contable_nombre=cuenta.nombre if cuenta else None,
        saldo_inicial=obj.saldo_inicial,
        activo=obj.activo,
    )


# ---------------------------------------------------------------------------
# Bancos
# ---------------------------------------------------------------------------

def listar_bancos(db: Session, solo_activos: bool = False) -> list[BanBanco]:
    q = db.query(BanBanco)
    if solo_activos:
        q = q.filter(BanBanco.activo == True)
    return q.order_by(BanBanco.nombre).all()


def crear_banco(db: Session, data: BancoCreate, actor: UsuarioActual) -> BanBanco:
    obj = BanBanco(**data.model_dump(), creado_por=uuid.UUID(actor.id))
    db.add(obj); db.commit(); db.refresh(obj)
    return obj


def actualizar_banco(db: Session, id: uuid.UUID, data: BancoUpdate, actor: UsuarioActual) -> BanBanco:
    obj = db.query(BanBanco).filter(BanBanco.id == id).first()
    if not obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Banco no encontrado")
    for campo, valor in data.model_dump(exclude_unset=True).items():
        setattr(obj, campo, valor)
    obj.modificado_por = uuid.UUID(actor.id)
    obj.modificado_en = datetime.now(timezone.utc)
    db.commit(); db.refresh(obj)
    return obj


# ---------------------------------------------------------------------------
# Cuentas bancarias
# ---------------------------------------------------------------------------

def listar_cuentas(db: Session, solo_activas: bool = False) -> list[CuentaBancariaResponse]:
    q = db.query(BanCuenta)
    if solo_activas:
        q = q.filter(BanCuenta.activo == True)
    return [_enrich_cuenta(db, o) for o in q.order_by(BanCuenta.nombre).all()]


def crear_cuenta(db: Session, data: CuentaBancariaCreate, actor: UsuarioActual) -> CuentaBancariaResponse:
    if not db.query(BanBanco).filter(BanBanco.id == data.banco_id).first():
        raise HTTPException(status_code=400, detail="Banco no encontrado")
    obj = BanCuenta(**data.model_dump(), creado_por=uuid.UUID(actor.id))
    db.add(obj); db.commit(); db.refresh(obj)
    return _enrich_cuenta(db, obj)


def actualizar_cuenta(db: Session, id: uuid.UUID, data: CuentaBancariaUpdate, actor: UsuarioActual) -> CuentaBancariaResponse:
    obj = db.query(BanCuenta).filter(BanCuenta.id == id).first()
    if not obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cuenta no encontrada")
    for campo, valor in data.model_dump(exclude_unset=True).items():
        setattr(obj, campo, valor)
    obj.modificado_por = uuid.UUID(actor.id)
    obj.modificado_en = datetime.now(timezone.utc)
    db.commit(); db.refresh(obj)
    return _enrich_cuenta(db, obj)


# ---------------------------------------------------------------------------
# Chequeras
# ---------------------------------------------------------------------------

def _enrich_chequera(db: Session, obj: BanChequera) -> ChequeraResponse:
    cuenta = db.query(BanCuenta).filter(BanCuenta.id == obj.cuenta_id).first()
    banco = db.query(BanBanco).filter(BanBanco.id == cuenta.banco_id).first() if cuenta else None
    return ChequeraResponse(
        id=obj.id, cuenta_id=obj.cuenta_id,
        cuenta_nombre=f"{cuenta.nombre} ({cuenta.numero})" if cuenta else None,
        banco_nombre=banco.nombre if banco else None,
        prefijo=obj.prefijo, numero_desde=obj.numero_desde,
        numero_hasta=obj.numero_hasta, consecutivo_actual=obj.consecutivo_actual,
        estado=obj.estado, descripcion=obj.descripcion, activo=obj.activo,
    )


def listar_chequeras(db: Session, solo_activas: bool = False) -> list[ChequeraResponse]:
    q = db.query(BanChequera)
    if solo_activas:
        q = q.filter(BanChequera.activo == True)
    return [_enrich_chequera(db, o) for o in q.order_by(BanChequera.numero_desde).all()]


def crear_chequera(db: Session, data: ChequerapCreate, actor: UsuarioActual) -> ChequeraResponse:
    if not db.query(BanCuenta).filter(BanCuenta.id == data.cuenta_id).first():
        raise HTTPException(status_code=400, detail="Cuenta bancaria no encontrada")
    if data.numero_hasta < data.numero_desde:
        raise HTTPException(status_code=400, detail="numero_hasta debe ser >= numero_desde")
    obj = BanChequera(
        **data.model_dump(),
        consecutivo_actual=data.numero_desde,
        estado="ACTIVA",
        creado_por=uuid.UUID(actor.id),
    )
    db.add(obj); db.commit(); db.refresh(obj)
    return _enrich_chequera(db, obj)


def actualizar_chequera(db: Session, id: uuid.UUID, data: ChequeraUpdate, actor: UsuarioActual) -> ChequeraResponse:
    obj = db.query(BanChequera).filter(BanChequera.id == id).first()
    if not obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Chequera no encontrada")
    payload = data.model_dump(exclude_unset=True)
    nd = payload.get("numero_desde", obj.numero_desde)
    nh = payload.get("numero_hasta", obj.numero_hasta)
    if nh < nd:
        raise HTTPException(status_code=400, detail="numero_hasta debe ser >= numero_desde")
    for campo, valor in payload.items():
        setattr(obj, campo, valor)
    obj.modificado_por = uuid.UUID(actor.id)
    obj.modificado_en = datetime.now(timezone.utc)
    db.commit(); db.refresh(obj)
    return _enrich_chequera(db, obj)


def movimientos_cuenta(db: Session, cuenta_id: uuid.UUID, fecha_desde: str | None = None, fecha_hasta: str | None = None):
    """Libro de bancos: movimientos contables de la cuenta bancaria con saldo corrido."""
    from decimal import Decimal
    from app.models.contabilidad import CntAsiento, CntAsientoLinea
    from app.schemas.bancos import MovimientosBancoResponse, MovimientoBancoItem

    cuenta = db.query(BanCuenta).filter(BanCuenta.id == cuenta_id).first()
    if not cuenta:
        raise HTTPException(status_code=404, detail="Cuenta bancaria no encontrada")
    cta_cont = db.get(CntCuenta, cuenta.cuenta_contable_id) if cuenta.cuenta_contable_id else None
    base = dict(
        cuenta_id=cuenta.id, cuenta_nombre=cuenta.nombre,
        cuenta_contable_codigo=cta_cont.codigo if cta_cont else None,
    )
    if not cuenta.cuenta_contable_id:
        return MovimientosBancoResponse(
            **base, saldo_inicial=Decimal("0"), saldo_final=Decimal("0"),
            total_debito=Decimal("0"), total_credito=Decimal("0"), items=[],
            aviso="La cuenta bancaria no tiene cuenta contable parametrizada; no hay movimientos que mostrar.",
        )

    q = (
        db.query(CntAsientoLinea, CntAsiento)
        .join(CntAsiento, CntAsientoLinea.asiento_id == CntAsiento.id)
        .filter(CntAsientoLinea.cuenta_id == cuenta.cuenta_contable_id, CntAsiento.estado == "publicado")
    )
    filas = q.all()

    # Saldo inicial = movimientos anteriores a fecha_desde
    saldo_inicial = Decimal("0")
    dentro = []
    for linea, asiento in filas:
        neto = linea.debito - linea.credito
        if fecha_desde and asiento.fecha < date.fromisoformat(fecha_desde):
            saldo_inicial += neto
        elif fecha_hasta and asiento.fecha > date.fromisoformat(fecha_hasta):
            continue
        else:
            dentro.append((asiento.fecha, asiento.numero, asiento.documento_numero, asiento.descripcion, linea.debito, linea.credito))

    dentro.sort(key=lambda x: (x[0], x[1]))
    saldo = saldo_inicial
    items = []
    tot_d = tot_c = Decimal("0")
    for fecha, numero, doc_num, desc, deb, cred in dentro:
        saldo += deb - cred
        tot_d += deb; tot_c += cred
        items.append(MovimientoBancoItem(
            fecha=fecha, asiento_numero=numero, documento_numero=doc_num,
            descripcion=desc, debito=deb, credito=cred, saldo=saldo,
        ))

    return MovimientosBancoResponse(
        **base, saldo_inicial=saldo_inicial, saldo_final=saldo,
        total_debito=tot_d, total_credito=tot_c, items=items,
    )


def movimientos_excel(db: Session, cuenta_id: uuid.UUID, fecha_desde: str | None = None, fecha_hasta: str | None = None):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = movimientos_cuenta(db, cuenta_id, fecha_desde, fecha_hasta)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro de bancos"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt = "#,##0.00"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Libro de bancos — {data.cuenta_nombre}" + (f" ({data.cuenta_contable_codigo})" if data.cuenta_contable_codigo else "")
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"Rango: {fecha_desde or 'inicio'} a {fecha_hasta or 'hoy'}"
    ws["A2"].font = Font(size=9, italic=True)

    headers = ["Fecha", "Documento", "Descripción", "Débito", "Crédito", "Saldo"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    fill = PatternFill(fill_type="solid", fgColor="EEEEEE")
    ws.cell(row=5, column=3, value="Saldo inicial").font = Font(italic=True, size=9)
    c = ws.cell(row=5, column=6, value=float(data.saldo_inicial)); c.number_format = num_fmt; c.font = Font(bold=True, size=9)

    row = 6
    for m in data.items:
        vals = [m.fecha.isoformat(), m.documento_numero or (f"#{m.asiento_numero}" if m.asiento_numero else ""),
                m.descripcion or "", float(m.debito), float(m.credito), float(m.saldo)]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = border
            if col >= 4:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")
        row += 1

    for col, v in [(3, "TOTALES"), (4, float(data.total_debito)), (5, float(data.total_credito)), (6, float(data.saldo_final))]:
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = Font(bold=True, size=10); cell.fill = fill; cell.border = border
        if col >= 4: cell.number_format = num_fmt; cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 44
    for col in ["D", "E", "F"]:
        ws.column_dimensions[col].width = 16

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"libro_bancos_{data.cuenta_nombre}_{fecha_hasta or 'hoy'}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )
