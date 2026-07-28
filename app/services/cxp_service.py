import uuid
from datetime import datetime, date, timezone
from decimal import Decimal

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.models.admin import AdmConcepto, AdmMoneda, AdmTipoDocumento, AdmConsecutivo
from app.models.adm import AdmTercero
from app.models.contabilidad import CntAsiento, CntAsientoLinea, CntCuenta, CntPeriodo
from app.models.bancos import BanCuenta
from app.models.cxp import CxpDocumento, CxpDocumentoLinea, CxpLineaRetencion, CxpParametroContable, CxpAplicacion
from app.schemas.auth import UsuarioActual
from app.schemas.cxp import (
    CxpDocumentoCreate, CxpDocumentoUpdate, AnularCxpRequest,
    CxpDocumentoResponse, CxpDocumentoListItem, CxpListResponse,
    CxpLineaResponse, LineaRetencionResponse,
    CxpResumenItem, CxpResumenResponse,
    ComprobanteCreate, FacturaPendienteCxpItem, AplicacionPendienteCxpItem,
)

TIPO_A_CODIGO = {
    "FACTURA":      "FCP",
    "COMPROBANTE":  "CP",
    "NOTA_CREDITO": "NCC",
    "NOTA_DEBITO":  "NDB",
    "ANTICIPO":     "ANTP",
    "VRT":          "VRT",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _buscar_periodo(db: Session, fecha: date) -> CntPeriodo:
    p = db.query(CntPeriodo).filter(
        CntPeriodo.fecha_inicio <= fecha,
        CntPeriodo.fecha_cierre >= fecha,
        CntPeriodo.activo == True,
    ).first()
    if not p:
        raise HTTPException(status_code=400, detail=f"No existe período contable para la fecha {fecha}")
    return p


def _moneda_funcional(db: Session) -> AdmMoneda:
    m = db.query(AdmMoneda).filter(AdmMoneda.es_funcional == True, AdmMoneda.activo == True).first()
    if not m:
        raise HTTPException(status_code=400, detail="No hay moneda funcional configurada")
    return m


def _get_tercero_nombre(db: Session, tercero_id: uuid.UUID) -> str:
    t = db.get(AdmTercero, tercero_id)
    return t.razon_social if t else ""


def _generar_numero(db: Session, tipo: str) -> str:
    codigo = TIPO_A_CODIGO.get(tipo, tipo)
    td = db.query(AdmTipoDocumento).filter(AdmTipoDocumento.codigo == codigo).first()
    if not td:
        raise HTTPException(status_code=400, detail=f"Tipo de documento {codigo} no configurado")
    cons = (
        db.query(AdmConsecutivo)
        .filter(AdmConsecutivo.tipo_documento_id == td.id)
        .with_for_update()
        .first()
    )
    if not cons:
        raise HTTPException(status_code=400, detail=f"No hay consecutivo configurado para {codigo}")
    siguiente = max(cons.numero_actual + 1, cons.numero_inicio)
    resultado = f"{cons.prefijo or ''}{str(siguiente).zfill(cons.longitud_minima)}"
    cons.numero_actual = siguiente
    return resultado


def _resolver_cuenta_gasto(db: Session, linea: CxpDocumentoLinea) -> CntCuenta | None:
    if linea.cuenta_id:
        return db.get(CntCuenta, linea.cuenta_id)
    if linea.concepto_id:
        concepto = db.get(AdmConcepto, linea.concepto_id)
        if concepto and concepto.cuenta_gasto_id:
            return db.get(CntCuenta, concepto.cuenta_gasto_id)
    return None


def _resolver_cuenta_cxp(db: Session, linea: CxpDocumentoLinea, fallback_id: uuid.UUID | None) -> CntCuenta | None:
    """Cuenta CxP (Proveedores) para acreditar en esta línea."""
    if linea.concepto_id:
        concepto = db.get(AdmConcepto, linea.concepto_id)
        if concepto and concepto.cuenta_cxp_id:
            return db.get(CntCuenta, concepto.cuenta_cxp_id)
    if fallback_id:
        return db.get(CntCuenta, fallback_id)
    return None


def _poblar_lineas_asiento(
    db: Session, asiento_id: uuid.UUID, doc: CxpDocumento,
    fallback_cxp_id: uuid.UUID | None, moneda_func: AdmMoneda,
) -> None:
    trm = doc.trm or Decimal("1")
    orden = 1

    def add_linea(cuenta_id, debito, credito, centro_costo_id=None):
        nonlocal orden
        d_func = (debito * trm).quantize(Decimal("0.0001")) if doc.moneda_id != moneda_func.id else debito
        c_func = (credito * trm).quantize(Decimal("0.0001")) if doc.moneda_id != moneda_func.id else credito
        db.add(CntAsientoLinea(
            id=uuid.uuid4(), asiento_id=asiento_id, orden=orden,
            cuenta_id=cuenta_id,
            debito=debito, credito=credito,
            debito_funcional=d_func, credito_funcional=c_func,
            tercero_id=doc.tercero_id,
            centro_costo_id=centro_costo_id,
        ))
        orden += 1

    # Nota crédito de proveedor: invierte los signos de la factura (reduce lo que debemos).
    rev = doc.tipo == "NOTA_CREDITO"

    def dc(deb, cred):
        return (cred, deb) if rev else (deb, cred)

    for linea in sorted(doc.lineas, key=lambda l: l.orden):
        cuenta_gasto = _resolver_cuenta_gasto(db, linea)
        if not cuenta_gasto:
            raise HTTPException(
                status_code=400,
                detail="No se pudo resolver la cuenta de gasto para una o más líneas"
            )
        cuenta_cxp = _resolver_cuenta_cxp(db, linea, fallback_cxp_id)
        if not cuenta_cxp:
            raise HTTPException(
                status_code=400,
                detail="No se pudo resolver la cuenta de proveedores para una o más líneas. "
                       "Verifica que el concepto tenga parametrizada la cuenta CxP, o configura "
                       "la cuenta global en Administración → Parámetros CxP."
            )

        add_linea(cuenta_gasto.id, *dc(linea.subtotal, Decimal("0")), linea.centro_costo_id)

        if linea.total_iva > 0 and linea.cuenta_iva_id:
            add_linea(linea.cuenta_iva_id, *dc(linea.total_iva, Decimal("0")))

        for ret in linea.retenciones:
            add_linea(ret.cuenta_id, *dc(Decimal("0"), ret.valor))

        # Proveedores = neto de esta línea (crédito en factura/ND, débito en NC)
        ret_linea = sum(r.valor for r in linea.retenciones)
        neto = linea.subtotal + linea.total_iva - ret_linea
        add_linea(cuenta_cxp.id, *dc(Decimal("0"), neto))


def _get_fallback_cxp_id(db: Session) -> uuid.UUID | None:
    params = db.query(CxpParametroContable).first()
    return params.cuenta_proveedores_id if params else None


def _cuenta_anticipos_cxp(db: Session) -> CntCuenta | None:
    params = db.query(CxpParametroContable).first()
    return db.get(CntCuenta, params.cuenta_anticipos_id) if params and params.cuenta_anticipos_id else None


def _cuenta_descuentos_cxp(db: Session) -> CntCuenta | None:
    params = db.query(CxpParametroContable).first()
    return db.get(CntCuenta, params.cuenta_descuentos_id) if params and params.cuenta_descuentos_id else None


def _cuenta_aprovechamientos_cxp(db: Session) -> CntCuenta | None:
    params = db.query(CxpParametroContable).first()
    return db.get(CntCuenta, params.cuenta_aprovechamientos_id) if params and params.cuenta_aprovechamientos_id else None


def _anticipos_total_comprobante(db: Session, comprobante_id) -> Decimal:
    """Σ de anticipos consumidos por un comprobante (credito=anticipo, debito=comprobante)."""
    if not comprobante_id:
        return Decimal("0")
    total = Decimal("0")
    apps = db.query(CxpAplicacion).filter(CxpAplicacion.documento_debito_id == comprobante_id).all()
    for ap in apps:
        cred = db.get(CxpDocumento, ap.documento_credito_id)
        if cred and cred.tipo == "ANTICIPO":
            total += ap.valor
    return total


def preview_asiento_documento(db: Session, data) -> "PreviewAsientoResponse":
    """Previsualiza el asiento de una factura/nota de CxP desde el payload (sin guardar)."""
    from app.schemas.facturacion import PreviewAsientoResponse, PreviewAsientoLinea
    from app.models.cxp import CxpLineaRetencion as _Ret
    avisos: list[str] = []
    fallback = _get_fallback_cxp_id(db)
    tercero_nombre = _get_tercero_nombre(db, data.tercero_id)
    lineas_out: list[PreviewAsientoLinea] = []

    def add(cuenta: CntCuenta | None, debito, credito, cc_id=None):
        cc_txt = None
        if cc_id:
            from app.models.contabilidad import CntCentroCosto
            cc = db.get(CntCentroCosto, cc_id)
            cc_txt = f"{cc.codigo} {cc.nombre}" if cc else None
        lineas_out.append(PreviewAsientoLinea(
            cuenta_codigo=cuenta.codigo if cuenta else None,
            cuenta_nombre=cuenta.nombre if cuenta else "(sin cuenta)",
            tercero_nombre=tercero_nombre, centro_costo=cc_txt,
            debito=debito, credito=credito,
        ))

    # ANTICIPO: Dr Anticipos a proveedores / Cr Banco (sin líneas de gasto)
    if getattr(data, "tipo", None) == "ANTICIPO":
        valor = getattr(data, "valor", None) or Decimal("0")
        cuenta_ant = _cuenta_anticipos_cxp(db)
        if not cuenta_ant:
            avisos.append("Falta la cuenta de Anticipos a proveedores en Parámetros CxP.")
        ban = db.get(BanCuenta, data.ban_cuenta_id) if getattr(data, "ban_cuenta_id", None) else None
        banco_cta = db.get(CntCuenta, ban.cuenta_contable_id) if ban and ban.cuenta_contable_id else None
        add(cuenta_ant, valor, Decimal("0"))
        add(banco_cta, Decimal("0"), valor)
        total_d = sum((l.debito for l in lineas_out), Decimal("0"))
        total_c = sum((l.credito for l in lineas_out), Decimal("0"))
        moneda = db.get(AdmMoneda, data.moneda_id)
        return PreviewAsientoResponse(
            lineas=lineas_out, total_debito=total_d, total_credito=total_c,
            cuadra=abs(total_d - total_c) <= Decimal("0.01"),
            moneda_codigo=moneda.codigo if moneda else None, avisos=avisos,
        )

    for ld in sorted(data.lineas, key=lambda l: l.orden):
        # Línea transitoria para reutilizar los resolvers.
        tmp = CxpDocumentoLinea(
            concepto_id=ld.concepto_id, cuenta_id=ld.cuenta_id,
            subtotal=ld.subtotal, total_iva=ld.total_iva,
            cuenta_iva_id=ld.cuenta_iva_id, centro_costo_id=ld.centro_costo_id,
        )
        cuenta_gasto = _resolver_cuenta_gasto(db, tmp)
        if not cuenta_gasto:
            avisos.append(f"Línea '{ld.descripcion}': sin cuenta de gasto (revisa el concepto o la cuenta).")
        cuenta_cxp = _resolver_cuenta_cxp(db, tmp, fallback)
        if not cuenta_cxp:
            avisos.append(f"Línea '{ld.descripcion}': sin cuenta de proveedores (concepto o Parámetros CxP).")
        add(cuenta_gasto, ld.subtotal, Decimal("0"), ld.centro_costo_id)
        if ld.total_iva > 0 and ld.cuenta_iva_id:
            add(db.get(CntCuenta, ld.cuenta_iva_id), ld.total_iva, Decimal("0"))
        ret_total = Decimal("0")
        for r in getattr(ld, "retenciones", []) or []:
            add(db.get(CntCuenta, r.cuenta_id) if r.cuenta_id else None, Decimal("0"), r.valor)
            ret_total += r.valor
        neto = ld.subtotal + ld.total_iva - ret_total
        add(cuenta_cxp, Decimal("0"), neto)

    total_d = sum((l.debito for l in lineas_out), Decimal("0"))
    total_c = sum((l.credito for l in lineas_out), Decimal("0"))
    moneda = db.get(AdmMoneda, data.moneda_id)
    return PreviewAsientoResponse(
        lineas=lineas_out, total_debito=total_d, total_credito=total_c,
        cuadra=abs(total_d - total_c) <= Decimal("0.01"),
        moneda_codigo=moneda.codigo if moneda else None, avisos=avisos,
    )


def _generar_asiento(db: Session, doc: CxpDocumento, actor: UsuarioActual) -> CntAsiento | None:
    moneda_func = _moneda_funcional(db)
    td_codigo = TIPO_A_CODIGO.get(doc.tipo, doc.tipo)
    td = db.query(AdmTipoDocumento).filter(AdmTipoDocumento.codigo == td_codigo).first()

    asiento = CntAsiento(
        id=uuid.uuid4(),
        tipo_documento_id=td.id if td else None,
        documento_numero=doc.numero,
        fecha=doc.fecha,
        periodo_id=doc.periodo_id,
        descripcion=f"{doc.tipo} {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}",
        estado="borrador",
        moneda_id=doc.moneda_id,
        trm=doc.trm if doc.moneda_id != moneda_func.id else None,
        documento_origen_id=doc.id,
        documento_origen_tipo="cxp_documento",
        creado_por=uuid.UUID(actor.id),
    )
    db.add(asiento)
    db.flush()
    _poblar_lineas_cxp(db, asiento.id, doc, moneda_func)
    return asiento


def _poblar_lineas_cxp(db: Session, asiento_id: uuid.UUID, doc: CxpDocumento, moneda_func: AdmMoneda) -> None:
    """Genera las líneas del asiento de un documento CxP según su tipo."""
    if doc.tipo == "ANTICIPO":
        cuenta_ant = _cuenta_anticipos_cxp(db)
        ban_cuenta = db.get(BanCuenta, doc.ban_cuenta_id) if doc.ban_cuenta_id else None
        if not cuenta_ant or not ban_cuenta or not ban_cuenta.cuenta_contable_id:
            raise HTTPException(status_code=400, detail="Configura la cuenta de anticipos en Parámetros CxP y la cuenta contable de la cuenta bancaria.")
        trm = doc.trm or Decimal("1")
        for cta_id, deb, cred in [
            (cuenta_ant.id, doc.total, Decimal("0")),
            (ban_cuenta.cuenta_contable_id, Decimal("0"), doc.total),
        ]:
            d_f = (deb * trm).quantize(Decimal("0.0001")) if doc.moneda_id != moneda_func.id else deb
            c_f = (cred * trm).quantize(Decimal("0.0001")) if doc.moneda_id != moneda_func.id else cred
            db.add(CntAsientoLinea(
                id=uuid.uuid4(), asiento_id=asiento_id, orden=1,
                cuenta_id=cta_id, debito=deb, credito=cred,
                debito_funcional=d_f, credito_funcional=c_f,
                tercero_id=doc.tercero_id,
            ))
        return
    _poblar_lineas_asiento(db, asiento_id, doc, _get_fallback_cxp_id(db), moneda_func)


def _calcular_totales_doc(doc: CxpDocumento) -> tuple[Decimal, Decimal, Decimal]:
    subtotal = sum(l.subtotal for l in doc.lineas)
    total_iva = sum(l.total_iva for l in doc.lineas)
    total_ret = sum(r.valor for l in doc.lineas for r in l.retenciones)
    return subtotal, total_iva, total_ret


def _to_linea_response(db: Session, linea: CxpDocumentoLinea) -> CxpLineaResponse:
    concepto_nombre = None
    if linea.concepto_id:
        c = db.get(AdmConcepto, linea.concepto_id)
        concepto_nombre = c.nombre if c else None

    cuenta_codigo = cuenta_nombre = None
    if linea.cuenta_id:
        c = db.get(CntCuenta, linea.cuenta_id)
        if c:
            cuenta_codigo, cuenta_nombre = c.codigo, c.nombre

    cuenta_iva_codigo = None
    if linea.cuenta_iva_id:
        c = db.get(CntCuenta, linea.cuenta_iva_id)
        cuenta_iva_codigo = c.codigo if c else None

    rets = []
    for r in linea.retenciones:
        c = db.get(CntCuenta, r.cuenta_id)
        rets.append(LineaRetencionResponse(
            id=r.id, tipo=r.tipo, descripcion=r.descripcion,
            base=r.base, porcentaje=r.porcentaje, valor=r.valor,
            cuenta_id=r.cuenta_id,
            cuenta_codigo=c.codigo if c else None,
            cuenta_nombre=c.nombre if c else None,
        ))

    return CxpLineaResponse(
        id=linea.id, orden=linea.orden, descripcion=linea.descripcion,
        concepto_id=linea.concepto_id, concepto_nombre=concepto_nombre,
        cuenta_id=linea.cuenta_id, cuenta_codigo=cuenta_codigo, cuenta_nombre=cuenta_nombre,
        subtotal=linea.subtotal, iva_pct=linea.iva_pct,
        total_iva=linea.total_iva, total=linea.total,
        centro_costo_id=linea.centro_costo_id,
        iva_tipo=linea.iva_tipo,
        cuenta_iva_id=linea.cuenta_iva_id, cuenta_iva_codigo=cuenta_iva_codigo,
        retenciones=rets,
    )


def _to_response(doc: CxpDocumento, db: Session) -> CxpDocumentoResponse:
    tercero = db.get(AdmTercero, doc.tercero_id)
    moneda = db.get(AdmMoneda, doc.moneda_id)
    return CxpDocumentoResponse(
        id=doc.id, numero=doc.numero, tipo=doc.tipo,
        numero_proveedor=doc.numero_proveedor,
        fecha=doc.fecha, fecha_vencimiento=doc.fecha_vencimiento,
        periodo_id=doc.periodo_id,
        tercero_id=doc.tercero_id,
        tercero_nit=tercero.nit if tercero else None,
        tercero_nombre=tercero.razon_social if tercero else None,
        moneda_id=doc.moneda_id,
        moneda_codigo=moneda.codigo if moneda else "",
        trm=doc.trm,
        subtotal=doc.subtotal, total_iva=doc.total_iva,
        total_retenciones=doc.total_retenciones,
        total=doc.total, saldo=doc.saldo,
        descripcion=doc.descripcion,
        condicion_pago_id=doc.condicion_pago_id,
        ban_cuenta_id=doc.ban_cuenta_id,
        estado=doc.estado,
        asiento_id=doc.asiento_id,
        asiento_modificado_manual=doc.asiento_modificado_manual,
        documento_origen_id=doc.documento_origen_id,
        factura_afectada_id=doc.factura_afectada_id,
        factura_afectada_numero=(db.get(CxpDocumento, doc.factura_afectada_id).numero if doc.factura_afectada_id else None),
        lineas=[_to_linea_response(db, l) for l in doc.lineas],
        creado_en=doc.creado_en,
        creado_por=doc.creado_por,
    )


def _to_list_item(doc: CxpDocumento, db: Session, hoy: date) -> CxpDocumentoListItem:
    tercero = db.get(AdmTercero, doc.tercero_id)
    moneda = db.get(AdmMoneda, doc.moneda_id)
    dias = None
    if doc.fecha_vencimiento and doc.estado == "contabilizado" and doc.saldo > 0:
        dias = (doc.fecha_vencimiento - hoy).days
    return CxpDocumentoListItem(
        id=doc.id, numero=doc.numero, tipo=doc.tipo,
        numero_proveedor=doc.numero_proveedor,
        fecha=doc.fecha, fecha_vencimiento=doc.fecha_vencimiento,
        tercero_nit=tercero.nit if tercero else None,
        tercero_nombre=tercero.razon_social if tercero else None,
        moneda_codigo=moneda.codigo if moneda else "",
        total=doc.total, saldo=doc.saldo,
        estado=doc.estado,
        dias_vencimiento=dias,
        factura_afectada_numero=(db.get(CxpDocumento, doc.factura_afectada_id).numero if doc.factura_afectada_id else None),
    )


def _persistir_lineas(db: Session, doc_id: uuid.UUID, lineas_data: list) -> None:
    for i, ld in enumerate(lineas_data, start=1):
        linea = CxpDocumentoLinea(
            id=uuid.uuid4(), documento_id=doc_id, orden=i,
            descripcion=ld.descripcion,
            concepto_id=ld.concepto_id, cuenta_id=ld.cuenta_id,
            subtotal=ld.subtotal, iva_pct=ld.iva_pct,
            total_iva=ld.total_iva, total=ld.total,
            centro_costo_id=ld.centro_costo_id,
            iva_tipo=f"GRAVADO_{int(ld.iva_pct)}" if ld.iva_tipo == "GRAVADO" else ld.iva_tipo,
            cuenta_iva_id=ld.cuenta_iva_id,
        )
        db.add(linea)
        db.flush()
        for ret in ld.retenciones:
            db.add(CxpLineaRetencion(
                id=uuid.uuid4(), linea_id=linea.id,
                tipo=ret.tipo, descripcion=ret.descripcion,
                base=ret.base, porcentaje=ret.porcentaje,
                valor=ret.valor, cuenta_id=ret.cuenta_id,
            ))


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------

def listar(
    db: Session,
    pagina: int = 1,
    por_pagina: int = 50,
    tipo: str | None = None,
    estado: str | None = None,
    tercero_id: uuid.UUID | None = None,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    solo_pendientes: bool = False,
) -> CxpListResponse:
    q = db.query(CxpDocumento).filter(CxpDocumento.activo == True)
    if tipo:          q = q.filter(CxpDocumento.tipo == tipo)
    if estado:        q = q.filter(CxpDocumento.estado == estado)
    if tercero_id:    q = q.filter(CxpDocumento.tercero_id == tercero_id)
    if fecha_desde:   q = q.filter(CxpDocumento.fecha >= fecha_desde)
    if fecha_hasta:   q = q.filter(CxpDocumento.fecha <= fecha_hasta)
    if solo_pendientes:
        q = q.filter(CxpDocumento.saldo > 0, CxpDocumento.estado == "contabilizado")

    total = q.count()
    hoy = date.today()
    rows = q.order_by(CxpDocumento.fecha.desc(), CxpDocumento.creado_en.desc()).offset((pagina - 1) * por_pagina).limit(por_pagina).all()
    return CxpListResponse(
        items=[_to_list_item(r, db, hoy) for r in rows],
        total=total, pagina=pagina, por_pagina=por_pagina,
    )


def obtener(db: Session, id: uuid.UUID) -> CxpDocumentoResponse:
    doc = db.query(CxpDocumento).filter(CxpDocumento.id == id, CxpDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return _to_response(doc, db)


def crear(db: Session, data: CxpDocumentoCreate, actor: UsuarioActual) -> CxpDocumentoResponse:
    periodo = _buscar_periodo(db, data.fecha)
    moneda_func = _moneda_funcional(db)

    if data.moneda_id != moneda_func.id and not data.trm:
        raise HTTPException(status_code=400, detail="Se requiere TRM para moneda extranjera")

    numero = _generar_numero(db, data.tipo)

    if data.tipo == "ANTICIPO":
        subtotal = data.valor or Decimal("0")
        total_iva = Decimal("0")
        total_ret = Decimal("0")
        total = subtotal
    else:
        subtotal = sum(l.subtotal for l in data.lineas)
        total_iva = sum(l.total_iva for l in data.lineas)
        total_ret = sum(r.valor for l in data.lineas for r in l.retenciones)
        total = subtotal + total_iva - total_ret

    doc = CxpDocumento(
        id=uuid.uuid4(), numero=numero, tipo=data.tipo,
        numero_proveedor=data.numero_proveedor,
        fecha=data.fecha, fecha_vencimiento=data.fecha_vencimiento,
        condicion_pago_id=data.condicion_pago_id,
        periodo_id=periodo.id, tercero_id=data.tercero_id,
        moneda_id=data.moneda_id,
        trm=data.trm if data.moneda_id != moneda_func.id else None,
        subtotal=subtotal, total_iva=total_iva,
        total_retenciones=total_ret, total=total, saldo=total,
        descripcion=data.descripcion,
        estado="borrador",
        ban_cuenta_id=data.ban_cuenta_id,
        factura_afectada_id=data.factura_afectada_id,
        creado_por=uuid.UUID(actor.id),
    )
    db.add(doc)
    db.flush()

    if data.tipo != "ANTICIPO":
        _persistir_lineas(db, doc.id, data.lineas)
    db.flush()
    db.refresh(doc)

    asiento = _generar_asiento(db, doc, actor)
    if asiento:
        doc.asiento_id = asiento.id

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def facturas_pendientes_cxp(
    db: Session,
    tercero_id: uuid.UUID,
    excluir_comprobante_id: uuid.UUID | None = None,
) -> list[FacturaPendienteCxpItem]:
    from sqlalchemy import func as sqlfunc
    hoy = date.today()
    docs = (
        db.query(CxpDocumento)
        .filter(
            CxpDocumento.activo == True,
            CxpDocumento.tercero_id == tercero_id,
            CxpDocumento.tipo.in_(["FACTURA", "VRT", "NOTA_DEBITO"]),
            CxpDocumento.estado == "contabilizado",
            CxpDocumento.saldo > 0,
        )
        .order_by(CxpDocumento.fecha_vencimiento.asc())
        .all()
    )
    result = []
    for d in docs:
        q = db.query(sqlfunc.coalesce(sqlfunc.sum(CxpAplicacion.valor), Decimal("0"))).filter(
            CxpAplicacion.documento_debito_id == d.id,
            CxpAplicacion.estado == "pendiente",
        )
        if excluir_comprobante_id:
            q = q.filter(CxpAplicacion.documento_credito_id != excluir_comprobante_id)
        pendiente = q.scalar()
        saldo_disp = d.saldo - pendiente
        if saldo_disp <= 0:
            continue
        dias = (d.fecha_vencimiento - hoy).days if d.fecha_vencimiento else None
        result.append(FacturaPendienteCxpItem(
            id=d.id, numero=d.numero, fecha=d.fecha,
            fecha_vencimiento=d.fecha_vencimiento,
            total=d.total, aplicado=(d.total - d.saldo) + pendiente,
            saldo=saldo_disp, dias_vencimiento=dias, tipo=d.tipo,
        ))
    return result


def _saldo_anticipo_disponible_cxp(db: Session, anticipo: CxpDocumento, excluir_comprobante_id) -> Decimal:
    from sqlalchemy import func
    q = db.query(func.coalesce(func.sum(CxpAplicacion.valor), Decimal("0"))).filter(
        CxpAplicacion.documento_credito_id == anticipo.id,
        CxpAplicacion.estado == "pendiente",
    )
    if excluir_comprobante_id:
        q = q.filter(CxpAplicacion.documento_debito_id != excluir_comprobante_id)
    return anticipo.saldo - q.scalar()


def anticipos_disponibles_cxp(db: Session, tercero_id: uuid.UUID, excluir_comprobante_id=None):
    from app.schemas.cxp import AnticipoDisponibleCxpItem
    docs = db.query(CxpDocumento).filter(
        CxpDocumento.activo == True,
        CxpDocumento.tercero_id == tercero_id,
        CxpDocumento.tipo == "ANTICIPO",
        CxpDocumento.estado == "contabilizado",
        CxpDocumento.saldo > 0,
    ).order_by(CxpDocumento.fecha.asc()).all()
    out = []
    for d in docs:
        disp = _saldo_anticipo_disponible_cxp(db, d, excluir_comprobante_id)
        if disp <= 0:
            continue
        out.append(AnticipoDisponibleCxpItem(id=d.id, numero=d.numero, fecha=d.fecha, total=d.total, saldo=disp))
    return out


def anticipos_aplicados_cxp(db: Session, comprobante_id: uuid.UUID):
    from app.schemas.cxp import AnticipoAplicadoCxpItem
    apps = db.query(CxpAplicacion).filter(
        CxpAplicacion.documento_debito_id == comprobante_id,
        CxpAplicacion.estado.in_(["pendiente", "aplicado"]),
    ).all()
    out = []
    for ap in apps:
        ant = db.get(CxpDocumento, ap.documento_credito_id)
        if not ant or ant.tipo != "ANTICIPO":
            continue
        out.append(AnticipoAplicadoCxpItem(anticipo_id=ant.id, numero=ant.numero, fecha=ant.fecha, valor=ap.valor))
    return out


def _validar_anticipos_comprobante(db: Session, tercero_id, anticipos, excluir_comprobante_id) -> Decimal:
    total = Decimal("0")
    for ant in anticipos:
        doc = db.query(CxpDocumento).filter(
            CxpDocumento.id == ant.anticipo_id,
            CxpDocumento.activo == True,
            CxpDocumento.tipo == "ANTICIPO",
            CxpDocumento.estado == "contabilizado",
        ).first()
        if not doc:
            raise HTTPException(status_code=400, detail="Anticipo no encontrado o no contabilizado")
        if doc.tercero_id != tercero_id:
            raise HTTPException(status_code=400, detail=f"El anticipo {doc.numero} no pertenece al proveedor")
        disp = _saldo_anticipo_disponible_cxp(db, doc, excluir_comprobante_id)
        if ant.valor > disp:
            raise HTTPException(status_code=400, detail=f"El valor del anticipo {doc.numero} ({ant.valor}) supera su saldo disponible ({disp})")
        total += ant.valor
    return total


def _validar_ajuste_comprobante(db: Session, ajuste: Decimal) -> None:
    if ajuste > 0 and not _cuenta_descuentos_cxp(db):
        raise HTTPException(status_code=400, detail="Configura la cuenta de Descuentos en Administración → Parámetros CxP para registrar la diferencia.")
    if ajuste < 0 and not _cuenta_aprovechamientos_cxp(db):
        raise HTTPException(status_code=400, detail="Configura la cuenta de Ajuste/pérdida en pagos en Administración → Parámetros CxP para registrar la diferencia.")


def notas_de_factura_cxp(db: Session, factura_id: uuid.UUID):
    from app.schemas.cxp import NotaRelacionadaCxpItem
    notas = db.query(CxpDocumento).filter(
        CxpDocumento.factura_afectada_id == factura_id,
        CxpDocumento.activo == True,
        CxpDocumento.tipo.in_(["NOTA_CREDITO", "NOTA_DEBITO"]),
    ).order_by(CxpDocumento.fecha.asc()).all()
    return [
        NotaRelacionadaCxpItem(
            id=n.id, numero=n.numero, tipo=n.tipo, fecha=n.fecha,
            total=n.total, saldo=n.saldo, estado=n.estado,
        )
        for n in notas
    ]


def cruces_de_documento_cxp(db: Session, doc_id: uuid.UUID):
    from app.schemas.cxp import CruceCxpItem
    apps = db.query(CxpAplicacion).filter(CxpAplicacion.documento_debito_id == doc_id).all()
    out = []
    for ap in apps:
        cred = db.get(CxpDocumento, ap.documento_credito_id)
        if not cred:
            continue
        out.append(CruceCxpItem(
            id=ap.id, documento_id=cred.id, numero=cred.numero, tipo=cred.tipo,
            fecha=ap.fecha, valor=ap.valor, estado=ap.estado,
        ))
    return out


def aplicar_cxp(db: Session, data, actor: UsuarioActual) -> dict:
    """Cruce anticipo/nota → factura de proveedor."""
    credito = db.query(CxpDocumento).filter(CxpDocumento.id == data.documento_credito_id, CxpDocumento.activo == True).first()
    debito = db.query(CxpDocumento).filter(CxpDocumento.id == data.documento_debito_id, CxpDocumento.activo == True).first()
    if not credito or not debito:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if credito.estado != "contabilizado" or debito.estado != "contabilizado":
        raise HTTPException(status_code=409, detail="Ambos documentos deben estar contabilizados")
    if credito.tercero_id != debito.tercero_id:
        raise HTTPException(status_code=400, detail="Los documentos deben pertenecer al mismo proveedor")
    if data.valor > credito.saldo:
        raise HTTPException(status_code=400, detail=f"El valor supera el saldo disponible del documento ({credito.saldo})")
    if data.valor > debito.saldo:
        raise HTTPException(status_code=400, detail=f"El valor supera el saldo de la factura ({debito.saldo})")

    db.add(CxpAplicacion(
        id=uuid.uuid4(), documento_credito_id=credito.id, documento_debito_id=debito.id,
        valor=data.valor, fecha=data.fecha, estado="aplicado", creado_por=uuid.UUID(actor.id),
    ))
    credito.saldo -= data.valor
    debito.saldo -= data.valor

    # Cruce de ANTICIPO con factura: reclasifica 1330 → Proveedores.
    if credito.tipo == "ANTICIPO":
        cuenta_ant = _cuenta_anticipos_cxp(db)
        cuenta_prov = db.get(CntCuenta, _get_fallback_cxp_id(db)) if _get_fallback_cxp_id(db) else None
        if not cuenta_ant or not cuenta_prov:
            raise HTTPException(status_code=400, detail="Faltan cuentas de anticipos/proveedores en Parámetros CxP para cruzar el anticipo.")
        periodo = _buscar_periodo(db, data.fecha)
        moneda_func = _moneda_funcional(db)
        trm = credito.trm or Decimal("1")
        asiento_ant = db.get(CntAsiento, credito.asiento_id) if credito.asiento_id else None
        asiento = CntAsiento(
            id=uuid.uuid4(),
            tipo_documento_id=asiento_ant.tipo_documento_id if asiento_ant else None,
            documento_numero=None,
            fecha=data.fecha, periodo_id=periodo.id,
            descripcion=f"CRUCE ANTICIPO {credito.numero} → {debito.numero}",
            estado="publicado", moneda_id=credito.moneda_id,
            trm=credito.trm if credito.moneda_id != moneda_func.id else None,
            documento_origen_id=credito.id, documento_origen_tipo="cxp_aplicacion",
            creado_por=uuid.UUID(actor.id),
        )
        db.add(asiento)
        db.flush()
        # Dr Proveedores / Cr Anticipos
        for orden, (cta_id, deb, cred) in enumerate(
            [(cuenta_prov.id, data.valor, Decimal("0")), (cuenta_ant.id, Decimal("0"), data.valor)], start=1
        ):
            d_f = (deb * trm).quantize(Decimal("0.0001")) if credito.moneda_id != moneda_func.id else deb
            c_f = (cred * trm).quantize(Decimal("0.0001")) if credito.moneda_id != moneda_func.id else cred
            db.add(CntAsientoLinea(
                id=uuid.uuid4(), asiento_id=asiento.id, orden=orden,
                cuenta_id=cta_id, debito=deb, credito=cred,
                debito_funcional=d_f, credito_funcional=c_f,
                tercero_id=credito.tercero_id,
            ))

    db.commit()
    return {"mensaje": f"Aplicación de {data.valor} registrada correctamente"}


def aplicaciones_comprobante(db: Session, comprobante_id: uuid.UUID) -> list[AplicacionPendienteCxpItem]:
    apps = db.query(CxpAplicacion).filter(
        CxpAplicacion.documento_credito_id == comprobante_id,
        CxpAplicacion.estado.in_(["pendiente", "aplicado"]),
    ).all()
    result = []
    for ap in apps:
        fac = db.get(CxpDocumento, ap.documento_debito_id)
        if not fac:
            continue
        # pendiente → fac.saldo no fue reducido, es el saldo real disponible
        # aplicado  → fac.saldo ya fue reducido, hay que sumar ap.valor para reconstruir el original
        saldo_original = fac.saldo if ap.estado == "pendiente" else fac.saldo + ap.valor
        result.append(AplicacionPendienteCxpItem(
            id=ap.id, factura_id=fac.id, numero=fac.numero,
            fecha=fac.fecha, fecha_vencimiento=fac.fecha_vencimiento,
            total=fac.total,
            saldo_original=saldo_original,
            valor=ap.valor,
        ))
    return result


def _cuenta_pago_documento(db: Session, debito_doc: CxpDocumento | None, cuenta_prov: CntCuenta) -> uuid.UUID:
    """Cuenta a debitar al pagar un documento: para VRT es su propia cuenta (2815),
    para el resto la cuenta de proveedores parametrizada."""
    if debito_doc and debito_doc.tipo == "VRT":
        linea = db.query(CxpDocumentoLinea).filter(
            CxpDocumentoLinea.documento_id == debito_doc.id,
            CxpDocumentoLinea.cuenta_id.isnot(None),
        ).first()
        if linea and linea.cuenta_id:
            return linea.cuenta_id
    return cuenta_prov.id


def preview_asiento_comprobante(db: Session, data) -> "PreviewAsientoResponse":
    """Previsualiza el asiento del comprobante de pago desde el payload (sin guardar)."""
    from app.schemas.facturacion import PreviewAsientoResponse, PreviewAsientoLinea
    avisos: list[str] = []
    fallback_cxp_id = _get_fallback_cxp_id(db)
    cuenta_prov = db.get(CntCuenta, fallback_cxp_id) if fallback_cxp_id else None
    if not cuenta_prov:
        avisos.append("No hay cuenta de proveedores parametrizada (Parámetros CxP).")
    ban_cuenta = db.get(BanCuenta, data.ban_cuenta_id) if data.ban_cuenta_id else None
    banco_cta = db.get(CntCuenta, ban_cuenta.cuenta_contable_id) if ban_cuenta and ban_cuenta.cuenta_contable_id else None
    if not banco_cta:
        avisos.append("La cuenta bancaria no tiene cuenta contable parametrizada.")

    tercero_nombre = _get_tercero_nombre(db, data.tercero_id)
    lineas_out = []

    # Débitos por documento aplicado (FACTURA → proveedores; VRT → su cuenta 2815).
    debitos: dict = {}
    for ap in data.aplicaciones:
        ddoc = db.get(CxpDocumento, ap.factura_id)
        cta_id = _cuenta_pago_documento(db, ddoc, cuenta_prov) if cuenta_prov else None
        debitos[cta_id] = debitos.get(cta_id, Decimal("0")) + ap.valor
    for cta_id, val in debitos.items():
        c = db.get(CntCuenta, cta_id) if cta_id else None
        lineas_out.append(PreviewAsientoLinea(
            cuenta_codigo=c.codigo if c else None,
            cuenta_nombre=c.nombre if c else "(sin cuenta)",
            tercero_nombre=tercero_nombre, centro_costo=None,
            debito=val, credito=Decimal("0"),
        ))
    # Crédito banco (efectivo)
    if data.valor_pagado > 0:
        lineas_out.append(PreviewAsientoLinea(
            cuenta_codigo=banco_cta.codigo if banco_cta else None,
            cuenta_nombre=banco_cta.nombre if banco_cta else "(sin cuenta banco)",
            tercero_nombre=tercero_nombre, centro_costo=None,
            debito=Decimal("0"), credito=data.valor_pagado,
        ))
    # Crédito anticipos (fuente)
    anticipos_total = sum((a.valor for a in getattr(data, "anticipos", [])), Decimal("0"))
    if anticipos_total > 0:
        cuenta_ant = _cuenta_anticipos_cxp(db)
        if not cuenta_ant:
            avisos.append("Falta la cuenta de Anticipos a proveedores en Parámetros CxP.")
        lineas_out.append(PreviewAsientoLinea(
            cuenta_codigo=cuenta_ant.codigo if cuenta_ant else None,
            cuenta_nombre=cuenta_ant.nombre if cuenta_ant else "(sin cuenta anticipos)",
            tercero_nombre=tercero_nombre, centro_costo=None,
            debito=Decimal("0"), credito=anticipos_total,
        ))
    # Ajuste = abonado − (efectivo + anticipos)
    total_aplicado = sum((a.valor for a in data.aplicaciones), Decimal("0"))
    ajuste = total_aplicado - data.valor_pagado - anticipos_total
    if ajuste > 0:
        cuenta_desc = _cuenta_descuentos_cxp(db)
        if not cuenta_desc:
            avisos.append("Falta la cuenta de Descuentos en Parámetros CxP.")
        lineas_out.append(PreviewAsientoLinea(
            cuenta_codigo=cuenta_desc.codigo if cuenta_desc else None,
            cuenta_nombre=cuenta_desc.nombre if cuenta_desc else "(sin cuenta descuentos)",
            tercero_nombre=tercero_nombre, centro_costo=None,
            debito=Decimal("0"), credito=ajuste,
        ))
    elif ajuste < 0:
        cuenta_apr = _cuenta_aprovechamientos_cxp(db)
        if not cuenta_apr:
            avisos.append("Falta la cuenta de Aprovechamientos en Parámetros CxP.")
        lineas_out.append(PreviewAsientoLinea(
            cuenta_codigo=cuenta_apr.codigo if cuenta_apr else None,
            cuenta_nombre=cuenta_apr.nombre if cuenta_apr else "(sin cuenta aprovechamientos)",
            tercero_nombre=tercero_nombre, centro_costo=None,
            debito=-ajuste, credito=Decimal("0"),
        ))

    total_d = sum((l.debito for l in lineas_out), Decimal("0"))
    total_c = sum((l.credito for l in lineas_out), Decimal("0"))
    moneda = db.get(AdmMoneda, data.moneda_id)
    return PreviewAsientoResponse(
        lineas=lineas_out, total_debito=total_d, total_credito=total_c,
        cuadra=abs(total_d - total_c) <= Decimal("0.01"),
        moneda_codigo=moneda.codigo if moneda else None, avisos=avisos,
    )


def asiento_contabilizado(db: Session, id: uuid.UUID) -> "PreviewAsientoResponse":
    """Devuelve las líneas del asiento REAL ya contabilizado de un documento CxP,
    en el mismo formato que el preview."""
    from app.schemas.facturacion import PreviewAsientoResponse, PreviewAsientoLinea
    doc = db.query(CxpDocumento).filter(CxpDocumento.id == id).first()
    if not doc or not doc.asiento_id:
        return PreviewAsientoResponse(
            lineas=[], total_debito=Decimal("0"), total_credito=Decimal("0"),
            cuadra=True, moneda_codigo=None,
            avisos=["El documento aún no tiene asiento contabilizado."],
        )
    asiento = db.get(CntAsiento, doc.asiento_id)
    lineas = db.query(CntAsientoLinea).filter(
        CntAsientoLinea.asiento_id == doc.asiento_id
    ).order_by(CntAsientoLinea.orden).all()
    out = []
    for l in lineas:
        c = db.get(CntCuenta, l.cuenta_id) if l.cuenta_id else None
        terc = db.get(AdmTercero, l.tercero_id) if l.tercero_id else None
        out.append(PreviewAsientoLinea(
            cuenta_codigo=c.codigo if c else None,
            cuenta_nombre=c.nombre if c else None,
            tercero_nombre=terc.razon_social if terc else None,
            centro_costo=None,
            debito=l.debito, credito=l.credito,
        ))
    total_d = sum((l.debito for l in lineas), Decimal("0"))
    total_c = sum((l.credito for l in lineas), Decimal("0"))
    moneda = db.get(AdmMoneda, doc.moneda_id)
    return PreviewAsientoResponse(
        lineas=out, total_debito=total_d, total_credito=total_c,
        cuadra=abs(total_d - total_c) <= Decimal("0.01"),
        moneda_codigo=moneda.codigo if moneda else None, avisos=[],
        asiento_numero=asiento.numero if asiento else None,
    )


def _generar_asiento_comprobante(db: Session, doc: CxpDocumento, actor: UsuarioActual) -> CntAsiento | None:
    fallback_cxp_id = _get_fallback_cxp_id(db)
    if not fallback_cxp_id:
        return None
    cuenta_prov = db.get(CntCuenta, fallback_cxp_id)
    if not cuenta_prov:
        return None
    if not doc.ban_cuenta_id:
        return None
    ban_cuenta = db.get(BanCuenta, doc.ban_cuenta_id)
    if not ban_cuenta or not ban_cuenta.cuenta_contable_id:
        return None

    moneda_func = _moneda_funcional(db)
    trm = doc.trm or Decimal("1")
    td = db.query(AdmTipoDocumento).filter(AdmTipoDocumento.codigo == "CP").first()

    asiento = CntAsiento(
        id=uuid.uuid4(),
        tipo_documento_id=td.id if td else None,
        documento_numero=doc.numero,
        fecha=doc.fecha,
        periodo_id=doc.periodo_id,
        descripcion=f"COMPROBANTE {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}",
        estado="borrador",
        moneda_id=doc.moneda_id,
        trm=doc.trm if doc.moneda_id != moneda_func.id else None,
        documento_origen_id=doc.id,
        documento_origen_tipo="cxp_documento",
        creado_por=uuid.UUID(actor.id),
    )
    db.add(asiento)
    db.flush()
    _poblar_lineas_comprobante(db, asiento.id, doc, moneda_func)
    return asiento


def _poblar_lineas_comprobante(db: Session, asiento_id: uuid.UUID, doc: CxpDocumento, moneda_func: AdmMoneda) -> None:
    """Líneas del asiento de un comprobante de egreso:
    Dr Proveedores (por doc aplicado) / Cr Banco (efectivo) + Cr Anticipos (fuente) + ajuste.
    """
    trm = doc.trm or Decimal("1")
    cuenta_prov = db.get(CntCuenta, _get_fallback_cxp_id(db)) if _get_fallback_cxp_id(db) else None
    ban_cuenta = db.get(BanCuenta, doc.ban_cuenta_id) if doc.ban_cuenta_id else None
    orden = 0

    def add(cuenta_id, debito, credito):
        nonlocal orden
        orden += 1
        d_f = (debito * trm).quantize(Decimal("0.0001")) if doc.moneda_id != moneda_func.id else debito
        c_f = (credito * trm).quantize(Decimal("0.0001")) if doc.moneda_id != moneda_func.id else credito
        db.add(CntAsientoLinea(
            id=uuid.uuid4(), asiento_id=asiento_id, orden=orden,
            cuenta_id=cuenta_id, debito=debito, credito=credito,
            debito_funcional=d_f, credito_funcional=c_f, tercero_id=doc.tercero_id,
        ))

    # Débitos: Proveedores por cada documento aplicado (VRT → su propia cuenta)
    apps = db.query(CxpAplicacion).filter(
        CxpAplicacion.documento_credito_id == doc.id,
        CxpAplicacion.estado.in_(["pendiente", "aplicado"]),
    ).all()
    debitos_por_cuenta: dict = {}
    total_aplicado = Decimal("0")
    for ap in apps:
        ddoc = db.get(CxpDocumento, ap.documento_debito_id)
        cta = _cuenta_pago_documento(db, ddoc, cuenta_prov)
        debitos_por_cuenta[cta] = debitos_por_cuenta.get(cta, Decimal("0")) + ap.valor
        total_aplicado += ap.valor
    if debitos_por_cuenta:
        for cta, val in debitos_por_cuenta.items():
            add(cta, val, Decimal("0"))
    else:
        add(cuenta_prov.id, doc.total, Decimal("0"))
        total_aplicado = doc.total

    efectivo = doc.subtotal
    anticipos_total = _anticipos_total_comprobante(db, doc.id)
    if efectivo > 0 and ban_cuenta and ban_cuenta.cuenta_contable_id:
        add(ban_cuenta.cuenta_contable_id, Decimal("0"), efectivo)
    if anticipos_total > 0:
        cuenta_ant = _cuenta_anticipos_cxp(db)
        if cuenta_ant:
            add(cuenta_ant.id, Decimal("0"), anticipos_total)

    # Ajuste = abonado a facturas − (efectivo + anticipos)
    ajuste = total_aplicado - efectivo - anticipos_total
    if ajuste > 0:            # el proveedor concedió un descuento → ingreso (Cr)
        cuenta_desc = _cuenta_descuentos_cxp(db)
        if cuenta_desc:
            add(cuenta_desc.id, Decimal("0"), ajuste)
    elif ajuste < 0:          # pagamos de más → aprovechamiento/gasto (Dr)
        cuenta_apr = _cuenta_aprovechamientos_cxp(db)
        if cuenta_apr:
            add(cuenta_apr.id, -ajuste, Decimal("0"))


def crear_comprobante(db: Session, data: ComprobanteCreate, actor: UsuarioActual) -> CxpDocumentoResponse:
    periodo = _buscar_periodo(db, data.fecha)
    moneda_func = _moneda_funcional(db)

    if data.moneda_id != moneda_func.id and not data.trm:
        raise HTTPException(status_code=400, detail="Se requiere TRM para moneda extranjera")

    ban_cuenta = db.get(BanCuenta, data.ban_cuenta_id)
    if not ban_cuenta or not ban_cuenta.activo:
        raise HTTPException(status_code=400, detail="Cuenta bancaria no encontrada")
    if not ban_cuenta.cuenta_contable_id:
        raise HTTPException(status_code=400, detail="La cuenta bancaria no tiene cuenta contable parametrizada")

    for ap in data.aplicaciones:
        fac = db.query(CxpDocumento).filter(
            CxpDocumento.id == ap.factura_id,
            CxpDocumento.activo == True,
            CxpDocumento.tipo.in_(["FACTURA", "VRT", "NOTA_DEBITO"]),
            CxpDocumento.estado == "contabilizado",
        ).first()
        if not fac:
            raise HTTPException(status_code=400, detail=f"Documento {ap.factura_id} no encontrado o no contabilizado")
        if fac.tercero_id != data.tercero_id:
            raise HTTPException(status_code=400, detail=f"El documento {fac.numero} no pertenece al tercero seleccionado")
        pendiente_otros = db.query(
            __import__("sqlalchemy").func.coalesce(
                __import__("sqlalchemy").func.sum(CxpAplicacion.valor), Decimal("0")
            )
        ).filter(
            CxpAplicacion.documento_debito_id == ap.factura_id,
            CxpAplicacion.estado == "pendiente",
        ).scalar()
        if ap.valor > fac.saldo - pendiente_otros:
            raise HTTPException(status_code=400, detail=f"El valor supera el saldo disponible de la factura {fac.numero}")

    total_anticipos = _validar_anticipos_comprobante(db, data.tercero_id, data.anticipos, excluir_comprobante_id=None)
    total_aplicado = sum((a.valor for a in data.aplicaciones), Decimal("0"))
    _validar_ajuste_comprobante(db, total_aplicado - (data.valor_pagado + total_anticipos))

    numero = _generar_numero(db, "COMPROBANTE")

    doc = CxpDocumento(
        id=uuid.uuid4(), numero=numero, tipo="COMPROBANTE",
        fecha=data.fecha, periodo_id=periodo.id,
        tercero_id=data.tercero_id, moneda_id=data.moneda_id,
        trm=data.trm if data.moneda_id != moneda_func.id else None,
        subtotal=data.valor_pagado,
        total_iva=Decimal("0"), total_retenciones=Decimal("0"),
        total=total_aplicado, saldo=total_aplicado,
        descripcion=data.descripcion,
        ban_cuenta_id=data.ban_cuenta_id,
        estado="borrador",
        creado_por=uuid.UUID(actor.id),
    )
    db.add(doc)
    db.flush()

    for ap in data.aplicaciones:
        db.add(CxpAplicacion(
            id=uuid.uuid4(),
            documento_credito_id=doc.id,
            documento_debito_id=ap.factura_id,
            valor=ap.valor, fecha=data.fecha,
            estado="pendiente",
            creado_por=uuid.UUID(actor.id),
        ))

    # Consumo de anticipos: credito = anticipo, debito = comprobante
    for ant in data.anticipos:
        db.add(CxpAplicacion(
            id=uuid.uuid4(),
            documento_credito_id=ant.anticipo_id,
            documento_debito_id=doc.id,
            valor=ant.valor, fecha=data.fecha,
            estado="pendiente",
            creado_por=uuid.UUID(actor.id),
        ))

    db.flush()
    db.refresh(doc)

    asiento = _generar_asiento_comprobante(db, doc, actor)
    if asiento:
        doc.asiento_id = asiento.id

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def actualizar_comprobante(db: Session, comprobante_id: uuid.UUID, data: ComprobanteCreate, actor: UsuarioActual) -> CxpDocumentoResponse:
    doc = db.query(CxpDocumento).filter(CxpDocumento.id == comprobante_id, CxpDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado")
    if doc.estado != "borrador":
        raise HTTPException(status_code=409, detail="Solo se pueden editar comprobantes en borrador")

    periodo = _buscar_periodo(db, data.fecha)
    moneda_func = _moneda_funcional(db)

    ban_cuenta = db.get(BanCuenta, data.ban_cuenta_id)
    if not ban_cuenta or not ban_cuenta.activo:
        raise HTTPException(status_code=400, detail="Cuenta bancaria no encontrada")
    if not ban_cuenta.cuenta_contable_id:
        raise HTTPException(status_code=400, detail="La cuenta bancaria no tiene cuenta contable parametrizada")

    for ap in data.aplicaciones:
        fac = db.query(CxpDocumento).filter(
            CxpDocumento.id == ap.factura_id,
            CxpDocumento.activo == True,
            CxpDocumento.tipo.in_(["FACTURA", "VRT", "NOTA_DEBITO"]),
            CxpDocumento.estado == "contabilizado",
        ).first()
        if not fac:
            raise HTTPException(status_code=400, detail=f"Documento {ap.factura_id} no encontrado")
        if fac.tercero_id != data.tercero_id:
            raise HTTPException(status_code=400, detail=f"El documento {fac.numero} no pertenece al tercero")
        app_actual = db.query(CxpAplicacion).filter(
            CxpAplicacion.documento_credito_id == comprobante_id,
            CxpAplicacion.documento_debito_id == ap.factura_id,
            CxpAplicacion.estado == "pendiente",
        ).first()
        saldo_disp = fac.saldo + (app_actual.valor if app_actual else Decimal("0"))
        if ap.valor > saldo_disp:
            raise HTTPException(status_code=400, detail=f"El valor supera el saldo de {fac.numero}")

    total_anticipos = _validar_anticipos_comprobante(db, data.tercero_id, data.anticipos, excluir_comprobante_id=comprobante_id)
    total_aplicado = sum((a.valor for a in data.aplicaciones), Decimal("0"))
    _validar_ajuste_comprobante(db, total_aplicado - (data.valor_pagado + total_anticipos))

    doc.fecha = data.fecha
    doc.periodo_id = periodo.id
    doc.tercero_id = data.tercero_id
    doc.moneda_id = data.moneda_id
    doc.trm = data.trm if data.moneda_id != moneda_func.id else None
    doc.subtotal = data.valor_pagado
    doc.total = total_aplicado
    doc.saldo = total_aplicado
    doc.descripcion = data.descripcion
    doc.ban_cuenta_id = data.ban_cuenta_id
    doc.modificado_por = uuid.UUID(actor.id)
    doc.modificado_en = datetime.now(timezone.utc)

    # Reemplazar aplicaciones (facturas: credito=comprobante) y consumos de anticipo (debito=comprobante)
    db.query(CxpAplicacion).filter(
        CxpAplicacion.documento_credito_id == comprobante_id,
        CxpAplicacion.estado == "pendiente",
    ).delete()
    db.query(CxpAplicacion).filter(
        CxpAplicacion.documento_debito_id == comprobante_id,
        CxpAplicacion.estado == "pendiente",
    ).delete()
    for ap in data.aplicaciones:
        db.add(CxpAplicacion(
            id=uuid.uuid4(),
            documento_credito_id=doc.id,
            documento_debito_id=ap.factura_id,
            valor=ap.valor, fecha=data.fecha,
            estado="pendiente",
            creado_por=uuid.UUID(actor.id),
        ))
    for ant in data.anticipos:
        db.add(CxpAplicacion(
            id=uuid.uuid4(),
            documento_credito_id=ant.anticipo_id,
            documento_debito_id=doc.id,
            valor=ant.valor, fecha=data.fecha,
            estado="pendiente",
            creado_por=uuid.UUID(actor.id),
        ))

    db.flush()
    db.refresh(doc)

    asiento = db.get(CntAsiento, doc.asiento_id) if doc.asiento_id else None
    if asiento and asiento.estado == "borrador":
        moneda_func2 = _moneda_funcional(db)
        asiento.fecha = doc.fecha
        asiento.periodo_id = doc.periodo_id
        asiento.descripcion = f"COMPROBANTE {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}"
        asiento.moneda_id = doc.moneda_id
        asiento.trm = doc.trm if doc.moneda_id != moneda_func2.id else None
        db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).delete()
        db.flush()
        _poblar_lineas_comprobante(db, asiento.id, doc, moneda_func2)
    else:
        nuevo = _generar_asiento_comprobante(db, doc, actor)
        if nuevo:
            doc.asiento_id = nuevo.id

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def listar_vrt(
    db: Session, tercero_id: uuid.UUID | None = None, estado: str | None = None,
    fecha_desde: str | None = None, fecha_hasta: str | None = None,
    pagina: int = 1, por_pagina: int = 50,
):
    from app.schemas.cxp import VrtItem, VrtListResponse
    from app.models.facturacion import FacFactura
    q = db.query(CxpDocumento).filter(CxpDocumento.activo == True, CxpDocumento.tipo == "VRT")
    if tercero_id:
        q = q.filter(CxpDocumento.tercero_id == tercero_id)
    if estado == "pendiente":
        q = q.filter(CxpDocumento.saldo > 0, CxpDocumento.estado == "contabilizado")
    elif estado == "pagado":
        q = q.filter(CxpDocumento.saldo <= 0, CxpDocumento.estado == "contabilizado")
    elif estado == "anulado":
        q = q.filter(CxpDocumento.estado == "anulado")
    if fecha_desde:
        q = q.filter(CxpDocumento.fecha >= fecha_desde)
    if fecha_hasta:
        q = q.filter(CxpDocumento.fecha <= fecha_hasta)
    total = q.count()
    rows = (q.order_by(CxpDocumento.fecha.desc(), CxpDocumento.creado_en.desc())
            .offset((pagina - 1) * por_pagina).limit(por_pagina).all())

    items = []
    for d in rows:
        tercero = db.get(AdmTercero, d.tercero_id)
        fac = db.get(FacFactura, d.origen_id) if d.origen_modulo == "fac_factura" and d.origen_id else None
        cliente = db.get(AdmTercero, fac.cliente_id) if fac else None
        if d.estado == "anulado":
            estado_pago = "anulado"
        elif d.saldo > 0:
            estado_pago = "pendiente"
        else:
            estado_pago = "pagado"
        comp_num, fecha_pago = None, None
        aps = db.query(CxpAplicacion).filter(
            CxpAplicacion.documento_debito_id == d.id,
            CxpAplicacion.estado == "aplicado",
        ).all()
        comps = [db.get(CxpDocumento, a.documento_credito_id) for a in aps]
        comps = [c for c in comps if c]
        if comps:
            comp_num = ", ".join(c.numero for c in comps)
            fecha_pago = max(c.fecha for c in comps)
        items.append(VrtItem(
            id=d.id, numero=d.numero, fecha=d.fecha,
            tercero_id=d.tercero_id,
            tercero_nit=tercero.nit if tercero else None,
            tercero_nombre=tercero.razon_social if tercero else None,
            valor=d.total, saldo=d.saldo, estado_pago=estado_pago,
            factura_id=fac.id if fac else None,
            factura_numero=fac.numero if fac else None,
            cliente_nombre=cliente.razon_social if cliente else None,
            comprobante_numero=comp_num, fecha_pago=fecha_pago,
        ))
    return VrtListResponse(items=items, total=total, pagina=pagina, por_pagina=por_pagina)


def resumen(db: Session, fecha_corte_str: str | None = None) -> CxpResumenResponse:
    from collections import defaultdict

    hoy = date.fromisoformat(fecha_corte_str) if fecha_corte_str else date.today()

    docs = db.query(CxpDocumento).filter(
        CxpDocumento.activo == True,
        CxpDocumento.estado == "contabilizado",
        CxpDocumento.saldo > 0,
        CxpDocumento.tipo.in_(["FACTURA", "VRT", "NOTA_DEBITO", "NOTA_CREDITO", "ANTICIPO"]),
    ).all()

    buckets: dict = defaultdict(lambda: {
        "corriente": Decimal("0"), "dias_1_30": Decimal("0"),
        "dias_31_60": Decimal("0"), "dias_61_90": Decimal("0"), "mas_90": Decimal("0"),
        "a_favor": Decimal("0"),
    })

    for doc in docs:
        b = buckets[doc.tercero_id]
        # NC y anticipos son saldo a favor (nos deben / prepagamos): columna aparte.
        if doc.tipo in ("NOTA_CREDITO", "ANTICIPO"):
            b["a_favor"] += doc.saldo
            continue
        if doc.fecha_vencimiento is None or doc.fecha_vencimiento >= hoy:
            b["corriente"] += doc.saldo
        else:
            dias = (hoy - doc.fecha_vencimiento).days
            if dias <= 30:   b["dias_1_30"]  += doc.saldo
            elif dias <= 60: b["dias_31_60"] += doc.saldo
            elif dias <= 90: b["dias_61_90"] += doc.saldo
            else:            b["mas_90"]     += doc.saldo

    items = []
    tot = {"corriente": Decimal("0"), "dias_1_30": Decimal("0"),
           "dias_31_60": Decimal("0"), "dias_61_90": Decimal("0"), "mas_90": Decimal("0"),
           "a_favor": Decimal("0")}

    for tercero_id, b in buckets.items():
        tercero = db.get(AdmTercero, tercero_id)
        total = (b["corriente"] + b["dias_1_30"] + b["dias_31_60"]
                 + b["dias_61_90"] + b["mas_90"] - b["a_favor"])
        items.append(CxpResumenItem(
            tercero_id=tercero_id,
            tercero_nit=tercero.nit if tercero else None,
            tercero_nombre=tercero.razon_social if tercero else None,
            corriente=b["corriente"], dias_1_30=b["dias_1_30"],
            dias_31_60=b["dias_31_60"], dias_61_90=b["dias_61_90"],
            mas_90=b["mas_90"], a_favor=b["a_favor"], total=total,
        ))
        for k in tot: tot[k] += b[k]

    items.sort(key=lambda x: x.mas_90 + x.dias_61_90 + x.dias_31_60, reverse=True)

    total_general = (tot["corriente"] + tot["dias_1_30"] + tot["dias_31_60"]
                     + tot["dias_61_90"] + tot["mas_90"] - tot["a_favor"])
    return CxpResumenResponse(
        fecha_corte=hoy, items=items,
        total_corriente=tot["corriente"], total_1_30=tot["dias_1_30"],
        total_31_60=tot["dias_31_60"], total_61_90=tot["dias_61_90"],
        total_mas_90=tot["mas_90"], total_a_favor=tot["a_favor"],
        total_general=total_general,
    )


def resumen_excel(db: Session, fecha_corte_str: str | None = None):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = resumen(db, fecha_corte_str)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saldos Proveedores"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=10)
    title_font  = Font(bold=True, size=12)
    num_fmt = '#,##0'

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Saldos de Proveedores — Fecha de corte: {data.fecha_corte}"
    ws["A1"].font = title_font

    headers = ["NIT", "Proveedor", "Corriente", "1 – 30 días", "31 – 60 días", "61 – 90 días", "+ 90 días", "A favor", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, item in enumerate(data.items, 4):
        vals = [
            item.tercero_nit or "",
            item.tercero_nombre or "",
            float(item.corriente),
            float(item.dias_1_30),
            float(item.dias_31_60),
            float(item.dias_61_90),
            float(item.mas_90),
            -float(item.a_favor),
            float(item.total),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            if col >= 3:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")

    total_row = len(data.items) + 4
    totals = ["TOTAL", "",
              float(data.total_corriente), float(data.total_1_30),
              float(data.total_31_60), float(data.total_61_90),
              float(data.total_mas_90), -float(data.total_a_favor), float(data.total_general)]
    fill = PatternFill(fill_type="solid", fgColor="EEEEEE")
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True, size=10)
        cell.fill = fill
        cell.border = border
        if col >= 3:
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"saldos_cxp_{data.fecha_corte}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def actualizar(db: Session, id: uuid.UUID, data: CxpDocumentoUpdate, actor: UsuarioActual) -> CxpDocumentoResponse:
    doc = db.query(CxpDocumento).filter(CxpDocumento.id == id, CxpDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.estado != "borrador":
        raise HTTPException(status_code=409, detail="Solo se pueden editar documentos en borrador")

    if data.fecha is not None:
        periodo = _buscar_periodo(db, data.fecha)
        doc.fecha = data.fecha
        doc.periodo_id = periodo.id
    if data.fecha_vencimiento is not None: doc.fecha_vencimiento = data.fecha_vencimiento
    if data.condicion_pago_id is not None: doc.condicion_pago_id = data.condicion_pago_id
    if data.numero_proveedor is not None:  doc.numero_proveedor = data.numero_proveedor
    if data.tercero_id is not None:        doc.tercero_id = data.tercero_id
    if data.moneda_id is not None:         doc.moneda_id = data.moneda_id
    if data.trm is not None:               doc.trm = data.trm
    if data.descripcion is not None:       doc.descripcion = data.descripcion
    if data.factura_afectada_id is not None: doc.factura_afectada_id = data.factura_afectada_id
    if data.ban_cuenta_id is not None:     doc.ban_cuenta_id = data.ban_cuenta_id

    if doc.tipo == "ANTICIPO":
        if data.valor is not None:
            doc.subtotal = data.valor
            doc.total_iva = Decimal("0")
            doc.total_retenciones = Decimal("0")
            doc.total = data.valor
            doc.saldo = data.valor
    elif data.lineas is not None:
        db.query(CxpDocumentoLinea).filter(CxpDocumentoLinea.documento_id == id).delete()
        db.flush()
        _persistir_lineas(db, doc.id, data.lineas)
        db.flush()
        db.refresh(doc)

        subtotal = sum(l.subtotal for l in doc.lineas)
        total_iva = sum(l.total_iva for l in doc.lineas)
        total_ret = sum(r.valor for l in doc.lineas for r in l.retenciones)
        total = subtotal + total_iva - total_ret
        doc.subtotal = subtotal
        doc.total_iva = total_iva
        doc.total_retenciones = total_ret
        doc.total = total
        doc.saldo = total

    doc.modificado_por = uuid.UUID(actor.id)
    doc.modificado_en = datetime.now(timezone.utc)

    # Regenerar asiento borrador
    if doc.asiento_id:
        asiento = db.get(CntAsiento, doc.asiento_id)
        if asiento and asiento.estado == "borrador":
            moneda_func = _moneda_funcional(db)
            asiento.fecha = doc.fecha
            asiento.periodo_id = doc.periodo_id
            asiento.descripcion = f"{doc.tipo} {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}"
            db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).delete()
            db.flush()
            _poblar_lineas_cxp(db, asiento.id, doc, moneda_func)
    else:
        asiento = _generar_asiento(db, doc, actor)
        if asiento:
            doc.asiento_id = asiento.id

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def contabilizar(db: Session, id: uuid.UUID, actor: UsuarioActual) -> CxpDocumentoResponse:
    doc = db.query(CxpDocumento).filter(CxpDocumento.id == id, CxpDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.estado != "borrador":
        raise HTTPException(status_code=409, detail="El documento ya está contabilizado o anulado")

    periodo = db.get(CntPeriodo, doc.periodo_id)
    if not periodo or periodo.estado != "abierto":
        raise HTTPException(status_code=400, detail="El período contable no está abierto")
    if doc.total <= 0:
        raise HTTPException(status_code=400, detail="El total del documento debe ser mayor que cero")
    if doc.tipo in ("FACTURA", "NOTA_DEBITO") and not doc.fecha_vencimiento:
        raise HTTPException(status_code=400, detail="La fecha de vencimiento es obligatoria")
    if doc.tipo in ("NOTA_CREDITO", "NOTA_DEBITO") and not doc.factura_afectada_id:
        raise HTTPException(status_code=400, detail="La nota debe referenciar la factura afectada antes de contabilizar")
    # COMPROBANTE: procesar aplicaciones pendientes
    if doc.tipo == "COMPROBANTE":
        apps = db.query(CxpAplicacion).filter(
            CxpAplicacion.documento_credito_id == id,
            CxpAplicacion.estado == "pendiente",
        ).all()
        if not apps:
            raise HTTPException(status_code=400, detail="El comprobante no tiene facturas aplicadas")
        for ap in apps:
            fac = db.query(CxpDocumento).filter(
                CxpDocumento.id == ap.documento_debito_id,
                CxpDocumento.activo == True,
            ).with_for_update().first()
            if not fac:
                raise HTTPException(status_code=400, detail="Factura aplicada no encontrada")
            if ap.valor > fac.saldo:
                raise HTTPException(status_code=400, detail=f"El saldo de {fac.numero} cambió. Revisa el comprobante.")
            fac.saldo -= ap.valor
            ap.estado = "aplicado"
        # Consumo de anticipos: reduce el saldo del anticipo (credito=anticipo, debito=comprobante)
        ant_apps = db.query(CxpAplicacion).filter(
            CxpAplicacion.documento_debito_id == id,
            CxpAplicacion.estado == "pendiente",
        ).all()
        for ap in ant_apps:
            ant = db.query(CxpDocumento).filter(
                CxpDocumento.id == ap.documento_credito_id,
                CxpDocumento.activo == True,
            ).with_for_update().first()
            if not ant or ant.tipo != "ANTICIPO":
                raise HTTPException(status_code=400, detail="Anticipo aplicado no encontrado")
            if ap.valor > ant.saldo:
                raise HTTPException(status_code=400, detail=f"El saldo del anticipo {ant.numero} cambió. Revisa el comprobante.")
            ant.saldo -= ap.valor
            ap.estado = "aplicado"
        doc.saldo = Decimal("0")

        # Generar/regenerar asiento comprobante
        if not doc.asiento_id:
            asiento_cp = _generar_asiento_comprobante(db, doc, actor)
            if not asiento_cp:
                raise HTTPException(
                    status_code=400,
                    detail="Configura la cuenta de proveedores en Parámetros CxP y la cuenta contable en la cuenta bancaria."
                )
            doc.asiento_id = asiento_cp.id
            db.flush()

        asiento = db.get(CntAsiento, doc.asiento_id)
        if not asiento or asiento.estado != "borrador":
            raise HTTPException(status_code=409, detail="El asiento ya está publicado o no fue encontrado")

        ban_cuenta = db.get(BanCuenta, doc.ban_cuenta_id)
        fallback = _get_fallback_cxp_id(db)
        cuenta_prov = db.get(CntCuenta, fallback) if fallback else None
        moneda_func = _moneda_funcional(db)
        if not cuenta_prov or not ban_cuenta or not ban_cuenta.cuenta_contable_id:
            raise HTTPException(status_code=400, detail="Parametriza cuenta proveedores y cuenta contable bancaria")

        trm = doc.trm or Decimal("1")
        asiento.fecha = doc.fecha
        asiento.periodo_id = doc.periodo_id
        asiento.descripcion = f"COMPROBANTE {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}"
        asiento.moneda_id = doc.moneda_id
        asiento.trm = doc.trm if doc.moneda_id != moneda_func.id else None
        db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).delete()
        db.flush()
        _poblar_lineas_comprobante(db, asiento.id, doc, moneda_func)
        db.flush()
        lineas = db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).all()
        total_d = sum(l.debito for l in lineas)
        total_c = sum(l.credito for l in lineas)
        if abs(total_d - total_c) > Decimal("0.01"):
            raise HTTPException(status_code=400, detail=f"Asiento descuadrado ({total_d} ≠ {total_c})")
        asiento.estado = "publicado"
        asiento.modificado_por = uuid.UUID(actor.id)
        asiento.modificado_en = datetime.now(timezone.utc)
        doc.estado = "contabilizado"
        doc.modificado_por = uuid.UUID(actor.id)
        doc.modificado_en = datetime.now(timezone.utc)
        db.commit()
        db.refresh(doc)
        return _to_response(doc, db)

    for linea in doc.lineas:
        for ret in linea.retenciones:
            if not ret.cuenta_id:
                raise HTTPException(
                    status_code=400,
                    detail=f"La retención '{ret.descripcion or ret.tipo}' no tiene cuenta contable asignada"
                )
        cuenta_gasto = _resolver_cuenta_gasto(db, linea)
        if cuenta_gasto:
            if cuenta_gasto.requiere_cc and not linea.centro_costo_id:
                raise HTTPException(
                    status_code=400,
                    detail=f"La cuenta {cuenta_gasto.codigo} — {cuenta_gasto.nombre} requiere centro de costo en la línea '{linea.descripcion or linea.orden}'"
                )
            if cuenta_gasto.requiere_tercero and not doc.tercero_id:
                raise HTTPException(
                    status_code=400,
                    detail=f"La cuenta {cuenta_gasto.codigo} — {cuenta_gasto.nombre} requiere tercero"
                )

    moneda_func = _moneda_funcional(db)
    fallback_cxp_id = _get_fallback_cxp_id(db)

    if not doc.asiento_id:
        asiento = _generar_asiento(db, doc, actor)
        if not asiento:
            raise HTTPException(status_code=400, detail="No se pudo generar el asiento contable")
        doc.asiento_id = asiento.id
        db.flush()

    asiento = db.get(CntAsiento, doc.asiento_id)
    if not asiento:
        raise HTTPException(status_code=400, detail="El asiento contable no fue encontrado")
    if asiento.estado != "borrador":
        raise HTTPException(status_code=409, detail="El asiento ya está publicado")

    # Regenerar líneas desde datos actuales
    asiento.fecha = doc.fecha
    asiento.periodo_id = doc.periodo_id
    asiento.descripcion = f"{doc.tipo} {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}"
    asiento.moneda_id = doc.moneda_id
    asiento.trm = doc.trm if doc.moneda_id != moneda_func.id else None
    db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).delete()
    db.flush()
    _poblar_lineas_cxp(db, asiento.id, doc, moneda_func)
    db.flush()

    lineas = db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).all()
    total_d = sum(l.debito for l in lineas)
    total_c = sum(l.credito for l in lineas)
    if abs(total_d - total_c) > Decimal("0.01"):
        raise HTTPException(
            status_code=400,
            detail=f"El asiento está descuadrado (débitos {total_d} ≠ créditos {total_c})"
        )

    asiento.estado = "publicado"
    asiento.modificado_por = uuid.UUID(actor.id)
    asiento.modificado_en = datetime.now(timezone.utc)

    doc.estado = "contabilizado"
    doc.saldo = doc.total
    doc.modificado_por = uuid.UUID(actor.id)
    doc.modificado_en = datetime.now(timezone.utc)

    # Nota crédito: cruce automático contra la factura afectada (reduce ambos saldos;
    # ambos viven en Proveedores → sin asiento adicional).
    if doc.tipo == "NOTA_CREDITO" and doc.factura_afectada_id:
        fac = db.query(CxpDocumento).filter(
            CxpDocumento.id == doc.factura_afectada_id, CxpDocumento.activo == True
        ).with_for_update().first()
        if fac and fac.estado == "contabilizado" and fac.saldo > 0:
            valor = min(doc.saldo, fac.saldo)
            if valor > 0:
                db.add(CxpAplicacion(
                    id=uuid.uuid4(), documento_credito_id=doc.id, documento_debito_id=fac.id,
                    valor=valor, fecha=doc.fecha, estado="aplicado", creado_por=uuid.UUID(actor.id),
                ))
                doc.saldo -= valor
                fac.saldo -= valor

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def anular(db: Session, id: uuid.UUID, data: AnularCxpRequest, actor: UsuarioActual) -> CxpDocumentoResponse:
    doc = db.query(CxpDocumento).filter(CxpDocumento.id == id, CxpDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.estado == "anulado":
        raise HTTPException(status_code=409, detail="El documento ya está anulado")

    # Verificar que este documento (como débito) no haya sido pagado por un comprobante.
    pagos = db.query(CxpAplicacion).filter(CxpAplicacion.documento_debito_id == id).all()
    for ap in pagos:
        cred = db.get(CxpDocumento, ap.documento_credito_id)
        if cred and cred.tipo == "COMPROBANTE" and cred.estado != "anulado":
            raise HTTPException(
                status_code=409,
                detail="El documento tiene comprobantes de pago aplicados. Anule o revierta el pago antes de anular."
            )

    # Si está contabilizada: verificar período y generar contraasiento
    if doc.estado == "contabilizado":
        periodo = db.get(CntPeriodo, doc.periodo_id)
        if not periodo or periodo.estado != "abierto":
            raise HTTPException(status_code=400, detail="El período contable no está abierto")

        if doc.asiento_id:
            asiento_orig = db.get(CntAsiento, doc.asiento_id)
            if asiento_orig:
                lineas_orig = db.query(CntAsientoLinea).filter(
                    CntAsientoLinea.asiento_id == asiento_orig.id
                ).all()

                contraasiento = CntAsiento(
                    id=uuid.uuid4(),
                    tipo_documento_id=asiento_orig.tipo_documento_id,
                    documento_numero=f"ANU-{asiento_orig.documento_numero}",
                    fecha=date.today(),
                    periodo_id=doc.periodo_id,
                    descripcion=f"ANULACIÓN {doc.tipo} {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)} · {data.motivo}",
                    estado="publicado",
                    moneda_id=asiento_orig.moneda_id,
                    trm=asiento_orig.trm,
                    documento_origen_id=doc.id,
                    documento_origen_tipo="cxp_documento",
                    creado_por=uuid.UUID(actor.id),
                )
                db.add(contraasiento)
                db.flush()

                for i, linea in enumerate(lineas_orig, start=1):
                    db.add(CntAsientoLinea(
                        id=uuid.uuid4(),
                        asiento_id=contraasiento.id,
                        orden=i,
                        cuenta_id=linea.cuenta_id,
                        debito=linea.credito,
                        credito=linea.debito,
                        debito_funcional=linea.credito_funcional,
                        credito_funcional=linea.debito_funcional,
                        tercero_id=linea.tercero_id,
                        centro_costo_id=linea.centro_costo_id,
                        descripcion=data.motivo,
                    ))

    elif doc.estado == "borrador" and doc.asiento_id:
        # Borrador: solo eliminar el asiento borrador
        asiento_borrador = db.get(CntAsiento, doc.asiento_id)
        if asiento_borrador and asiento_borrador.estado == "borrador":
            db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento_borrador.id).delete()
            db.delete(asiento_borrador)
            doc.asiento_id = None

    # Revertir aplicaciones donde este documento abona/paga a otros (credito = doc):
    # comprobante→factura, NC→factura, anticipo→factura. Restaura el saldo del débito.
    apps_credito = db.query(CxpAplicacion).filter(
        CxpAplicacion.documento_credito_id == doc.id,
        CxpAplicacion.estado.in_(["pendiente", "aplicado"]),
    ).all()
    for ap in apps_credito:
        if ap.estado == "aplicado":
            otro = db.query(CxpDocumento).filter(CxpDocumento.id == ap.documento_debito_id).with_for_update().first()
            if otro:
                otro.saldo += ap.valor
        db.delete(ap)

    # Revertir consumos de anticipo de un comprobante (credito = anticipo, debito = doc): restaura el anticipo.
    apps_debito = db.query(CxpAplicacion).filter(
        CxpAplicacion.documento_debito_id == doc.id,
        CxpAplicacion.estado.in_(["pendiente", "aplicado"]),
    ).all()
    for ap in apps_debito:
        cred = db.get(CxpDocumento, ap.documento_credito_id)
        if cred and cred.tipo == "ANTICIPO":
            if ap.estado == "aplicado":
                ant = db.query(CxpDocumento).filter(CxpDocumento.id == cred.id).with_for_update().first()
                if ant:
                    ant.saldo += ap.valor
            db.delete(ap)

    doc.estado = "anulado"
    doc.saldo = Decimal("0")
    doc.modificado_por = uuid.UUID(actor.id)
    doc.modificado_en = datetime.now(timezone.utc)
    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)
