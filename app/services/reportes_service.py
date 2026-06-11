from collections import defaultdict
from datetime import date
from decimal import Decimal
import uuid as _uuid

from sqlalchemy import func, case
from sqlalchemy.orm import Session

from app.models.adm import AdmTercero
from app.models.contabilidad import CntAsiento, CntAsientoLinea, CntCuenta


def balanza_comprobacion(
    db: Session,
    fecha_desde: date,
    fecha_hasta: date,
    nivel: int = 3,
) -> dict:
    CERO = Decimal("0")

    # Movimientos acumulados hasta fecha_hasta agrupados por cuenta
    rows = (
        db.query(
            CntAsientoLinea.cuenta_id,
            func.sum(
                case((CntAsiento.fecha < fecha_desde, CntAsientoLinea.debito_funcional), else_=CERO)
            ).label("si_debito"),
            func.sum(
                case((CntAsiento.fecha < fecha_desde, CntAsientoLinea.credito_funcional), else_=CERO)
            ).label("si_credito"),
            func.sum(
                case(
                    (
                        (CntAsiento.fecha >= fecha_desde) & (CntAsiento.fecha <= fecha_hasta),
                        CntAsientoLinea.debito_funcional,
                    ),
                    else_=CERO,
                )
            ).label("periodo_debito"),
            func.sum(
                case(
                    (
                        (CntAsiento.fecha >= fecha_desde) & (CntAsiento.fecha <= fecha_hasta),
                        CntAsientoLinea.credito_funcional,
                    ),
                    else_=CERO,
                )
            ).label("periodo_credito"),
        )
        .join(CntAsiento, CntAsientoLinea.asiento_id == CntAsiento.id)
        .filter(
            CntAsiento.estado == "publicado",
            CntAsiento.fecha <= fecha_hasta,
            CntAsientoLinea.activo == True,
        )
        .group_by(CntAsientoLinea.cuenta_id)
        .all()
    )

    # Índice cuenta_id → movimientos
    movs: dict = {r.cuenta_id: r for r in rows}

    # Cuentas al nivel solicitado que tienen movimientos propios o en descendientes
    cuentas = (
        db.query(CntCuenta)
        .filter(CntCuenta.nivel == nivel, CntCuenta.activo == True)
        .order_by(CntCuenta.codigo)
        .all()
    )

    # Para cuentas de nivel > hoja, necesitamos agregar hacia arriba.
    # Estrategia: para cada cuenta al nivel solicitado, sumar los movimientos
    # de todas las cuentas cuyo código empieza con el código de esa cuenta.
    cuenta_por_id = {c.id: c for c in db.query(CntCuenta).filter(CntCuenta.activo == True).all()}

    def _agregar(cuenta: CntCuenta) -> tuple:
        prefix = cuenta.codigo
        si_d = si_c = per_d = per_c = CERO
        for cid, r in movs.items():
            c = cuenta_por_id.get(cid)
            if c and c.codigo.startswith(prefix):
                si_d  += r.si_debito
                si_c  += r.si_credito
                per_d += r.periodo_debito
                per_c += r.periodo_credito
        return si_d, si_c, per_d, per_c

    filas = []
    tot_si_d = tot_si_c = tot_per_d = tot_per_c = tot_sf_d = tot_sf_c = CERO

    for cuenta in cuentas:
        si_d, si_c, per_d, per_c = _agregar(cuenta)

        # Omitir cuentas sin ningún movimiento
        if si_d == si_c == per_d == per_c == CERO:
            continue

        sf_d = si_d + per_d
        sf_c = si_c + per_c

        # Netos para presentación: si el saldo final neto es deudor, va en sf_debito; si acreedor, en sf_credito
        sf_neto = sf_d - sf_c
        sf_debito  = sf_neto if sf_neto > 0 else CERO
        sf_credito = (-sf_neto) if sf_neto < 0 else CERO

        si_neto = si_d - si_c
        si_debito  = si_neto if si_neto > 0 else CERO
        si_credito = (-si_neto) if si_neto < 0 else CERO

        tot_si_d  += si_debito
        tot_si_c  += si_credito
        tot_per_d += per_d
        tot_per_c += per_c
        tot_sf_d  += sf_debito
        tot_sf_c  += sf_credito

        filas.append({
            "codigo":      cuenta.codigo,
            "nombre":      cuenta.nombre,
            "naturaleza":  cuenta.naturaleza,
            "si_debito":   si_debito,
            "si_credito":  si_credito,
            "periodo_debito":  per_d,
            "periodo_credito": per_c,
            "sf_debito":   sf_debito,
            "sf_credito":  sf_credito,
        })

    cuadrado = (
        abs(tot_per_d - tot_per_c) < Decimal("0.01")
        and abs(tot_sf_d - tot_sf_c) < Decimal("0.01")
    )

    return {
        "fecha_desde": str(fecha_desde),
        "fecha_hasta": str(fecha_hasta),
        "nivel": nivel,
        "filas": filas,
        "totales": {
            "si_debito":       tot_si_d,
            "si_credito":      tot_si_c,
            "periodo_debito":  tot_per_d,
            "periodo_credito": tot_per_c,
            "sf_debito":       tot_sf_d,
            "sf_credito":      tot_sf_c,
        },
        "cuadrado": cuadrado,
    }


def _mayor_queries(db, cuenta_ids, fecha_desde, fecha_hasta, t_id):
    """Devuelve (si_map, mov_rows, terceros_map) para un conjunto de cuentas."""
    CERO = Decimal("0")

    si_q = (
        db.query(
            CntAsientoLinea.cuenta_id,
            func.sum(CntAsientoLinea.debito_funcional).label("deb"),
            func.sum(CntAsientoLinea.credito_funcional).label("cred"),
        )
        .join(CntAsiento, CntAsientoLinea.asiento_id == CntAsiento.id)
        .filter(
            CntAsiento.estado == "publicado",
            CntAsientoLinea.activo == True,
            CntAsientoLinea.cuenta_id.in_(cuenta_ids),
            CntAsiento.fecha < fecha_desde,
        )
    )
    if t_id:
        si_q = si_q.filter(CntAsientoLinea.tercero_id == t_id)
    si_rows = si_q.group_by(CntAsientoLinea.cuenta_id).all()
    si_map = {
        r.cuenta_id: (Decimal(str(r.deb or 0)), Decimal(str(r.cred or 0)))
        for r in si_rows
    }

    mov_q = (
        db.query(
            CntAsientoLinea.cuenta_id,
            CntAsientoLinea.tercero_id,
            CntAsientoLinea.descripcion.label("linea_desc"),
            CntAsientoLinea.debito_funcional,
            CntAsientoLinea.credito_funcional,
            CntAsientoLinea.orden,
            CntAsiento.fecha,
            CntAsiento.numero,
            CntAsiento.descripcion.label("asiento_desc"),
        )
        .join(CntAsiento, CntAsientoLinea.asiento_id == CntAsiento.id)
        .filter(
            CntAsiento.estado == "publicado",
            CntAsientoLinea.activo == True,
            CntAsientoLinea.cuenta_id.in_(cuenta_ids),
            CntAsiento.fecha >= fecha_desde,
            CntAsiento.fecha <= fecha_hasta,
        )
        .order_by(CntAsiento.fecha, CntAsiento.numero, CntAsientoLinea.orden)
    )
    if t_id:
        mov_q = mov_q.filter(CntAsientoLinea.tercero_id == t_id)
    mov_rows = mov_q.all()

    t_ids = {r.tercero_id for r in mov_rows if r.tercero_id}
    terceros_map: dict = {}
    if t_ids:
        for t in db.query(AdmTercero.id, AdmTercero.nit, AdmTercero.razon_social).filter(AdmTercero.id.in_(t_ids)):
            terceros_map[t.id] = {"nit": t.nit, "nombre": t.razon_social}

    return si_map, mov_rows, terceros_map


def libro_mayor(
    db: Session,
    cuenta_desde: str,
    cuenta_hasta: str,
    fecha_desde: date,
    fecha_hasta: date,
    tercero_id: str | None = None,
) -> dict:
    CERO = Decimal("0")

    cuentas = (
        db.query(CntCuenta)
        .filter(
            CntCuenta.codigo >= cuenta_desde,
            CntCuenta.codigo <= cuenta_hasta,
            CntCuenta.activo == True,
        )
        .order_by(CntCuenta.codigo)
        .all()
    )
    cuenta_ids = [c.id for c in cuentas]

    if not cuenta_ids:
        return {"fecha_desde": str(fecha_desde), "fecha_hasta": str(fecha_hasta),
                "cuenta_desde": cuenta_desde, "cuenta_hasta": cuenta_hasta, "cuentas": []}

    t_id = _uuid.UUID(tercero_id) if tercero_id else None
    si_map, mov_rows, terceros_map = _mayor_queries(db, cuenta_ids, fecha_desde, fecha_hasta, t_id)

    mov_by_cuenta: dict = defaultdict(list)
    for r in mov_rows:
        mov_by_cuenta[r.cuenta_id].append(r)

    result = []
    for cuenta in cuentas:
        si_deb, si_cred = si_map.get(cuenta.id, (CERO, CERO))
        si_neto = (si_deb - si_cred) if cuenta.naturaleza == "DEBITO" else (si_cred - si_deb)

        lineas = []
        saldo = si_neto
        tot_deb = tot_cred = CERO

        for r in mov_by_cuenta.get(cuenta.id, []):
            deb  = Decimal(str(r.debito_funcional  or 0))
            cred = Decimal(str(r.credito_funcional or 0))
            saldo += (deb - cred) if cuenta.naturaleza == "DEBITO" else (cred - deb)
            tot_deb  += deb
            tot_cred += cred
            t = terceros_map.get(r.tercero_id, {})
            lineas.append({
                "fecha":          str(r.fecha),
                "numero":         r.numero,
                "descripcion":    r.linea_desc or r.asiento_desc or "",
                "tercero_nit":    t.get("nit", ""),
                "tercero_nombre": t.get("nombre", ""),
                "debito":         str(deb),
                "credito":        str(cred),
                "saldo":          str(saldo),
            })

        if si_deb == si_cred == CERO and not lineas:
            continue

        result.append({
            "cuenta_codigo":  cuenta.codigo,
            "cuenta_nombre":  cuenta.nombre,
            "naturaleza":     cuenta.naturaleza,
            "saldo_inicial":  str(si_neto),
            "lineas":         lineas,
            "totales": {
                "debito":      str(tot_deb),
                "credito":     str(tot_cred),
                "saldo_final": str(saldo),
            },
        })

    return {
        "fecha_desde":  str(fecha_desde),
        "fecha_hasta":  str(fecha_hasta),
        "cuenta_desde": cuenta_desde,
        "cuenta_hasta": cuenta_hasta,
        "cuentas":      result,
    }


def libro_mayor_excel(
    db: Session,
    cuenta_desde: str,
    cuenta_hasta: str,
    fecha_desde: date,
    fecha_hasta: date,
    tercero_id: str | None = None,
):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = libro_mayor(db, cuenta_desde, cuenta_hasta, fecha_desde, fecha_hasta, tercero_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro Mayor"

    thin   = Side(style="thin", color="CCCCCC")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    num    = '#,##0.00'
    right  = Alignment(horizontal="right")
    center = Alignment(horizontal="center")
    fill_h = PatternFill(fill_type="solid", fgColor="1F4E79")
    fill_a = PatternFill(fill_type="solid", fgColor="DDEEFF")
    fill_t = PatternFill(fill_type="solid", fgColor="EEEEEE")

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Libro Mayor — {fecha_desde} a {fecha_hasta}"
    ws["A1"].font = Font(bold=True, size=13)

    HDR = ["Fecha", "N° Asiento", "Descripción", "Identificación", "Nombre Tercero", "Débito", "Crédito", "Saldo"]
    COLS = ["A", "B", "C", "D", "E", "F", "G", "H"]
    WIDTHS = [12, 12, 36, 16, 28, 16, 16, 16]

    row = 3
    for cuenta_data in data["cuentas"]:
        # Encabezado cuenta
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1,
                    value=f"{cuenta_data['cuenta_codigo']}  {cuenta_data['cuenta_nombre']}")
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = fill_h
        row += 1

        # Columnas
        for ci, h in enumerate(HDR, 1):
            cel = ws.cell(row=row, column=ci, value=h)
            cel.font = Font(bold=True, size=9)
            cel.fill = fill_a
            cel.border = brd
            cel.alignment = right if ci >= 5 else Alignment(horizontal="left")
        row += 1

        # Saldo inicial
        ws.cell(row=row, column=3, value="SALDO INICIAL").font = Font(italic=True)
        ws.cell(row=row, column=7, value=float(cuenta_data["saldo_inicial"])).number_format = num
        ws.cell(row=row, column=7).alignment = right
        row += 1

        # Líneas
        for li, ln in enumerate(cuenta_data["lineas"]):
            bg = None if li % 2 == 0 else PatternFill(fill_type="solid", fgColor="F7F7F7")
            vals = [ln["fecha"], ln["numero"], ln["descripcion"],
                    ln["tercero_nit"], ln["tercero_nombre"],
                    float(ln["debito"]) or None, float(ln["credito"]) or None, float(ln["saldo"])]
            for ci, val in enumerate(vals, 1):
                cel = ws.cell(row=row, column=ci, value=val)
                cel.border = brd
                if ci >= 6:
                    cel.number_format = num; cel.alignment = right
                if bg:
                    cel.fill = bg
            row += 1

        # Totales
        tots = cuenta_data["totales"]
        ws.cell(row=row, column=3, value="TOTALES").font = Font(bold=True)
        for ci, val in [(6, tots["debito"]), (7, tots["credito"]), (8, tots["saldo_final"])]:
            cel = ws.cell(row=row, column=ci, value=float(val))
            cel.font = Font(bold=True); cel.number_format = num
            cel.alignment = right; cel.fill = fill_t; cel.border = brd
        row += 2

    for col, w in zip(COLS, WIDTHS):
        ws.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fn = f"mayor_{cuenta_desde}_{cuenta_hasta}_{fecha_desde}_{fecha_hasta}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


def auxiliar_tercero(
    db: Session,
    cuenta_desde: str | None,
    cuenta_hasta: str | None,
    fecha_desde: date,
    fecha_hasta: date,
    tercero_id: str | None = None,
) -> dict:
    CERO = Decimal("0")

    q = db.query(CntCuenta).filter(CntCuenta.activo == True)
    if cuenta_desde:
        q = q.filter(CntCuenta.codigo >= cuenta_desde)
    if cuenta_hasta:
        q = q.filter(CntCuenta.codigo <= cuenta_hasta)
    cuentas = q.order_by(CntCuenta.codigo).all()
    cuenta_ids = [c.id for c in cuentas]

    if not cuenta_ids:
        return {"fecha_desde": str(fecha_desde), "fecha_hasta": str(fecha_hasta),
                "cuenta_desde": cuenta_desde, "cuenta_hasta": cuenta_hasta, "cuentas": []}

    t_id = _uuid.UUID(tercero_id) if tercero_id else None
    si_map_raw, mov_rows, terceros_map = _mayor_queries(db, cuenta_ids, fecha_desde, fecha_hasta, t_id)

    # Saldo inicial desagregado por (cuenta, tercero)
    si_q2 = (
        db.query(
            CntAsientoLinea.cuenta_id,
            CntAsientoLinea.tercero_id,
            func.sum(CntAsientoLinea.debito_funcional).label("deb"),
            func.sum(CntAsientoLinea.credito_funcional).label("cred"),
        )
        .join(CntAsiento, CntAsientoLinea.asiento_id == CntAsiento.id)
        .filter(
            CntAsiento.estado == "publicado",
            CntAsientoLinea.activo == True,
            CntAsientoLinea.cuenta_id.in_(cuenta_ids),
            CntAsiento.fecha < fecha_desde,
        )
    )
    if t_id:
        si_q2 = si_q2.filter(CntAsientoLinea.tercero_id == t_id)
    si_by_ct = {
        (r.cuenta_id, r.tercero_id): (Decimal(str(r.deb or 0)), Decimal(str(r.cred or 0)))
        for r in si_q2.group_by(CntAsientoLinea.cuenta_id, CntAsientoLinea.tercero_id).all()
    }

    # Cargar nombre de todos los terceros referenciados
    all_t_ids = {r.tercero_id for r in mov_rows if r.tercero_id} | {k[1] for k in si_by_ct if k[1]}
    all_terceros: dict = {}
    if all_t_ids:
        for t in db.query(AdmTercero.id, AdmTercero.nit, AdmTercero.razon_social).filter(AdmTercero.id.in_(all_t_ids)):
            all_terceros[t.id] = {"nit": t.nit, "nombre": t.razon_social}

    # Agrupar movs por (cuenta, tercero)
    mov_by_ct: dict = defaultdict(list)
    for r in mov_rows:
        mov_by_ct[(r.cuenta_id, r.tercero_id)].append(r)

    cuenta_map = {c.id: c for c in cuentas}
    result = []

    for cuenta in cuentas:
        # Recopilar todos los tercero_id que aparecen en esta cuenta
        tkeys = set()
        for (cid, tid) in si_by_ct:
            if cid == cuenta.id:
                tkeys.add(tid)
        for (cid, tid) in mov_by_ct:
            if cid == cuenta.id:
                tkeys.add(tid)

        if not tkeys:
            continue

        tercero_sections = []
        for tid in sorted(tkeys, key=lambda x: all_terceros.get(x, {}).get("nombre", "") if x else ""):
            si_deb, si_cred = si_by_ct.get((cuenta.id, tid), (CERO, CERO))
            si_neto = (si_deb - si_cred) if cuenta.naturaleza == "DEBITO" else (si_cred - si_deb)

            lineas = []
            saldo = si_neto
            tot_deb = tot_cred = CERO

            for r in mov_by_ct.get((cuenta.id, tid), []):
                deb  = Decimal(str(r.debito_funcional  or 0))
                cred = Decimal(str(r.credito_funcional or 0))
                saldo += (deb - cred) if cuenta.naturaleza == "DEBITO" else (cred - deb)
                tot_deb  += deb
                tot_cred += cred
                lineas.append({
                    "fecha":       str(r.fecha),
                    "numero":      r.numero,
                    "descripcion": r.linea_desc or r.asiento_desc or "",
                    "debito":      str(deb),
                    "credito":     str(cred),
                    "saldo":       str(saldo),
                })

            t_info = all_terceros.get(tid, {}) if tid else {}
            tercero_sections.append({
                "tercero_id":    str(tid) if tid else None,
                "tercero_nit":   t_info.get("nit", ""),
                "tercero_nombre": t_info.get("nombre", "Sin tercero") if tid else "Sin tercero",
                "saldo_inicial": str(si_neto),
                "lineas":        lineas,
                "totales": {
                    "debito":      str(tot_deb),
                    "credito":     str(tot_cred),
                    "saldo_final": str(saldo),
                },
            })

        result.append({
            "cuenta_codigo": cuenta.codigo,
            "cuenta_nombre": cuenta.nombre,
            "naturaleza":    cuenta.naturaleza,
            "terceros":      tercero_sections,
        })

    return {
        "fecha_desde":  str(fecha_desde),
        "fecha_hasta":  str(fecha_hasta),
        "cuenta_desde": cuenta_desde,
        "cuenta_hasta": cuenta_hasta,
        "cuentas":      result,
    }


def auxiliar_excel(
    db: Session,
    cuenta_desde: str | None,
    cuenta_hasta: str | None,
    fecha_desde: date,
    fecha_hasta: date,
    tercero_id: str | None = None,
):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = auxiliar_tercero(db, cuenta_desde, cuenta_hasta, fecha_desde, fecha_hasta, tercero_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auxiliar Terceros"

    thin  = Side(style="thin", color="CCCCCC")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    num   = '#,##0.00'
    right = Alignment(horizontal="right")
    fill_c = PatternFill(fill_type="solid", fgColor="1F4E79")
    fill_t = PatternFill(fill_type="solid", fgColor="2E75B6")
    fill_h = PatternFill(fill_type="solid", fgColor="DDEEFF")
    fill_s = PatternFill(fill_type="solid", fgColor="EEEEEE")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Auxiliar por Tercero — {fecha_desde} a {fecha_hasta}"
    ws["A1"].font = Font(bold=True, size=13)

    HDR = ["Fecha", "N° Asiento", "Descripción", "Débito", "Crédito", "Saldo"]

    row = 3
    for cuenta_data in data["cuentas"]:
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1,
                    value=f"{cuenta_data['cuenta_codigo']}  {cuenta_data['cuenta_nombre']}")
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = fill_c
        row += 1

        for sec in cuenta_data["terceros"]:
            ws.merge_cells(f"A{row}:F{row}")
            label_t = f"{sec['tercero_nit']}  {sec['tercero_nombre']}" if sec.get("tercero_nit") else sec["tercero_nombre"]
            c = ws.cell(row=row, column=1, value=label_t)
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = fill_t
            row += 1

            for ci, h in enumerate(HDR, 1):
                cel = ws.cell(row=row, column=ci, value=h)
                cel.font = Font(bold=True, size=9)
                cel.fill = fill_h; cel.border = brd
                cel.alignment = right if ci >= 4 else Alignment(horizontal="left")
            row += 1

            ws.cell(row=row, column=3, value="SALDO INICIAL").font = Font(italic=True)
            ws.cell(row=row, column=6, value=float(sec["saldo_inicial"])).number_format = num
            ws.cell(row=row, column=6).alignment = right
            row += 1

            for li, ln in enumerate(sec["lineas"]):
                bg = None if li % 2 == 0 else PatternFill(fill_type="solid", fgColor="F7F7F7")
                vals = [ln["fecha"], ln["numero"], ln["descripcion"],
                        float(ln["debito"]) or None, float(ln["credito"]) or None, float(ln["saldo"])]
                for ci, val in enumerate(vals, 1):
                    cel = ws.cell(row=row, column=ci, value=val)
                    cel.border = brd
                    if ci >= 4:
                        cel.number_format = num; cel.alignment = right
                    if bg:
                        cel.fill = bg
                row += 1

            tots = sec["totales"]
            ws.cell(row=row, column=3, value="TOTALES").font = Font(bold=True)
            for ci, val in [(4, tots["debito"]), (5, tots["credito"]), (6, tots["saldo_final"])]:
                cel = ws.cell(row=row, column=ci, value=float(val))
                cel.font = Font(bold=True); cel.number_format = num
                cel.alignment = right; cel.fill = fill_s; cel.border = brd
            row += 2

    for col, w in zip(["A","B","C","D","E","F"], [12, 12, 44, 16, 16, 16]):
        ws.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fn = f"auxiliar_{cuenta_desde}_{cuenta_hasta}_{fecha_desde}_{fecha_hasta}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


def estado_resultados(db: Session, fecha_desde: date, fecha_hasta: date, nivel: int = 3) -> dict:
    CERO = Decimal("0")

    rows = (
        db.query(
            CntAsientoLinea.cuenta_id,
            func.sum(CntAsientoLinea.debito_funcional).label("total_debito"),
            func.sum(CntAsientoLinea.credito_funcional).label("total_credito"),
        )
        .join(CntAsiento, CntAsientoLinea.asiento_id == CntAsiento.id)
        .filter(
            CntAsiento.estado == "publicado",
            CntAsiento.fecha >= fecha_desde,
            CntAsiento.fecha <= fecha_hasta,
            CntAsientoLinea.activo == True,
        )
        .group_by(CntAsientoLinea.cuenta_id)
        .all()
    )

    movs: dict = {r.cuenta_id: r for r in rows}
    cuenta_por_id = {c.id: c for c in db.query(CntCuenta).filter(CntCuenta.activo == True).all()}
    cuentas_nivel = (
        db.query(CntCuenta)
        .filter(CntCuenta.nivel <= nivel, CntCuenta.activo == True)
        .order_by(CntCuenta.codigo)
        .all()
    )

    def _agregar(cuenta: CntCuenta) -> tuple:
        prefix = cuenta.codigo
        deb = cred = CERO
        for cid, r in movs.items():
            c = cuenta_por_id.get(cid)
            if c and c.codigo.startswith(prefix):
                deb  += r.total_debito
                cred += r.total_credito
        return deb, cred

    def _saldo_neto(cuenta: CntCuenta, deb: Decimal, cred: Decimal) -> Decimal:
        if cuenta.naturaleza == "DEBITO":
            return deb - cred
        return cred - deb

    def _clasificar(codigo: str) -> str | None:
        c1 = codigo[0]
        c2 = codigo[:2]
        if c1 == "4":
            return "ing_oper" if c2 == "41" else "ing_no_oper"
        if c1 in ("6", "7"):
            return "costos"
        if c1 == "5":
            return "gasto_oper" if c2 in ("51", "52") else "gasto_no_oper"
        return None

    buckets: dict[str, list] = {
        "ing_oper": [], "ing_no_oper": [],
        "costos": [],
        "gasto_oper": [], "gasto_no_oper": [],
    }

    for cuenta in cuentas_nivel:
        sec = _clasificar(cuenta.codigo)
        if sec is None:
            continue
        deb, cred = _agregar(cuenta)
        if deb == cred == CERO:
            continue
        saldo = _saldo_neto(cuenta, deb, cred)
        if saldo == CERO:
            continue
        buckets[sec].append({"codigo": cuenta.codigo, "nombre": cuenta.nombre,
                              "saldo": saldo, "nivel": cuenta.nivel})

    def _total(sec: str) -> Decimal:
        return sum(i["saldo"] for i in buckets[sec] if i["nivel"] == nivel)

    tot_ing_oper     = _total("ing_oper")
    tot_costos       = _total("costos")
    tot_gasto_oper   = _total("gasto_oper")
    tot_ing_no_oper  = _total("ing_no_oper")
    tot_gasto_no_oper= _total("gasto_no_oper")

    util_bruta            = tot_ing_oper - tot_costos
    util_operacional      = util_bruta - tot_gasto_oper
    util_antes_impuestos  = util_operacional + tot_ing_no_oper - tot_gasto_no_oper
    util_neta             = util_antes_impuestos  # impuesto de renta no separado aún

    def _ser(items: list) -> list:
        return [{"codigo": i["codigo"], "nombre": i["nombre"],
                 "saldo": str(i["saldo"]), "nivel": i["nivel"]} for i in items]

    return {
        "fecha_desde": str(fecha_desde),
        "fecha_hasta": str(fecha_hasta),
        "nivel": nivel,
        "ing_oper":       _ser(buckets["ing_oper"]),
        "costos":         _ser(buckets["costos"]),
        "gasto_oper":     _ser(buckets["gasto_oper"]),
        "ing_no_oper":    _ser(buckets["ing_no_oper"]),
        "gasto_no_oper":  _ser(buckets["gasto_no_oper"]),
        "subtotales": {
            "ing_oper":             str(tot_ing_oper),
            "costos":               str(tot_costos),
            "utilidad_bruta":       str(util_bruta),
            "gasto_oper":           str(tot_gasto_oper),
            "utilidad_operacional": str(util_operacional),
            "ing_no_oper":          str(tot_ing_no_oper),
            "gasto_no_oper":        str(tot_gasto_no_oper),
            "utilidad_antes_impuestos": str(util_antes_impuestos),
            "utilidad_neta":        str(util_neta),
        },
    }


def resultados_excel(db: Session, fecha_desde: date, fecha_hasta: date, nivel: int = 3):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = estado_resultados(db, fecha_desde, fecha_hasta, nivel)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num    = '#,##0.00'
    right  = Alignment(horizontal="right")
    fill_sub = PatternFill(fill_type="solid", fgColor="EEEEEE")
    fill_util = PatternFill(fill_type="solid", fgColor="D6EAF8")

    ws.merge_cells("A1:C1")
    ws["A1"] = f"Estado de Resultados — {fecha_desde} a {fecha_hasta}"
    ws["A1"].font = Font(bold=True, size=12)

    row = 3

    SECCIONES = [
        ("INGRESOS OPERACIONALES",   data["ing_oper"],     data["subtotales"]["ing_oper"],    False),
        ("COSTO DE VENTAS",          data["costos"],       data["subtotales"]["costos"],       False),
        ("GASTOS OPERACIONALES",     data["gasto_oper"],   data["subtotales"]["gasto_oper"],   False),
        ("INGRESOS NO OPERACIONALES",data["ing_no_oper"],  data["subtotales"]["ing_no_oper"],  False),
        ("GASTOS NO OPERACIONALES",  data["gasto_no_oper"],data["subtotales"]["gasto_no_oper"],False),
    ]
    SUBTOTALES = [
        ("= UTILIDAD BRUTA",              data["subtotales"]["utilidad_bruta"],           True),
        ("= UTILIDAD OPERACIONAL",        data["subtotales"]["utilidad_operacional"],      True),
        ("= UTILIDAD ANTES IMPUESTOS",    data["subtotales"]["utilidad_antes_impuestos"],  True),
        ("= UTILIDAD NETA",               data["subtotales"]["utilidad_neta"],             True),
    ]
    sub_iter = iter(SUBTOTALES)

    for label, items, total_sec, _ in SECCIONES:
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, size=10)
        row += 1
        for item in items:
            indent = (item["nivel"] - 1) * 2
            ws.cell(row=row, column=1, value=item["codigo"]).border = border
            cn = ws.cell(row=row, column=2, value=(" " * indent) + item["nombre"])
            cn.border = border
            if item["nivel"] < nivel:
                cn.font = Font(bold=True, size=9.5)
            c3 = ws.cell(row=row, column=3, value=float(item["saldo"]))
            c3.number_format = num; c3.alignment = right; c3.border = border
            row += 1
        # Total de sección
        ct = ws.cell(row=row, column=2, value=f"Total {label}")
        ct.font = Font(bold=True); ct.fill = fill_sub; ct.border = border
        ws.cell(row=row, column=1, value="").fill = fill_sub
        c3 = ws.cell(row=row, column=3, value=float(total_sec))
        c3.font = Font(bold=True); c3.number_format = num; c3.alignment = right
        c3.fill = fill_sub; c3.border = border
        row += 1
        # Subtotal de utilidad después de ciertos grupos
        if label in ("COSTO DE VENTAS", "GASTOS OPERACIONALES",
                     "GASTOS NO OPERACIONALES"):
            try:
                sub_label, sub_val, _ = next(sub_iter)
            except StopIteration:
                continue
            for col in [1, 2, 3]:
                ws.cell(row=row, column=col).fill = fill_util
                ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=2, value=sub_label).font = Font(bold=True, size=11)
            c3 = ws.cell(row=row, column=3, value=float(sub_val))
            c3.font = Font(bold=True, size=11); c3.number_format = num; c3.alignment = right
            row += 2

    # UTILIDAD NETA final
    for col in [1, 2, 3]:
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=2, value="= UTILIDAD NETA").font = Font(bold=True, size=12, color="FFFFFF")
    c3 = ws.cell(row=row, column=3, value=float(data["subtotales"]["utilidad_neta"]))
    c3.font = Font(bold=True, size=12, color="FFFFFF"); c3.number_format = num; c3.alignment = right

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 20

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"resultados_{fecha_desde}_{fecha_hasta}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def balance_general(db: Session, fecha_corte: date, nivel: int = 3) -> dict:
    CERO = Decimal("0")

    rows = (
        db.query(
            CntAsientoLinea.cuenta_id,
            func.sum(CntAsientoLinea.debito_funcional).label("total_debito"),
            func.sum(CntAsientoLinea.credito_funcional).label("total_credito"),
        )
        .join(CntAsiento, CntAsientoLinea.asiento_id == CntAsiento.id)
        .filter(
            CntAsiento.estado == "publicado",
            CntAsiento.fecha <= fecha_corte,
            CntAsientoLinea.activo == True,
        )
        .group_by(CntAsientoLinea.cuenta_id)
        .all()
    )

    movs: dict = {r.cuenta_id: r for r in rows}
    cuenta_por_id = {c.id: c for c in db.query(CntCuenta).filter(CntCuenta.activo == True).all()}
    # Incluye todos los niveles 1..nivel para mostrar jerarquía completa
    cuentas_nivel = (
        db.query(CntCuenta)
        .filter(CntCuenta.nivel <= nivel, CntCuenta.activo == True)
        .order_by(CntCuenta.codigo)
        .all()
    )

    def _agregar(cuenta: CntCuenta) -> tuple:
        prefix = cuenta.codigo
        deb = cred = CERO
        for cid, r in movs.items():
            c = cuenta_por_id.get(cid)
            if c and c.codigo.startswith(prefix):
                deb  += r.total_debito
                cred += r.total_credito
        return deb, cred

    def _saldo_neto(cuenta: CntCuenta, deb: Decimal, cred: Decimal) -> Decimal:
        if cuenta.naturaleza == "DEBITO":
            return deb - cred
        return cred - deb

    activo: list = []
    pasivo: list = []
    patrimonio: list = []
    ingresos_neto = CERO
    gastos_neto   = CERO

    for cuenta in cuentas_nivel:
        deb, cred = _agregar(cuenta)
        if deb == cred == CERO:
            continue
        saldo = _saldo_neto(cuenta, deb, cred)
        if saldo == CERO:
            continue
        clase = cuenta.codigo[0]
        item = {"codigo": cuenta.codigo, "nombre": cuenta.nombre, "saldo": saldo, "nivel": cuenta.nivel}
        if clase == "1":
            activo.append(item)
        elif clase == "2":
            pasivo.append(item)
        elif clase == "3":
            patrimonio.append(item)
        elif clase == "4" and cuenta.nivel == nivel:
            ingresos_neto += saldo
        elif clase in ("5", "6", "7") and cuenta.nivel == nivel:
            gastos_neto += saldo

    utilidad = ingresos_neto - gastos_neto
    # Solo cuentas hoja para evitar doble conteo de niveles padre + hijo
    total_activo    = sum(i["saldo"] for i in activo    if i["nivel"] == nivel)
    total_pasivo    = sum(i["saldo"] for i in pasivo    if i["nivel"] == nivel)
    total_patrimonio_cuentas = sum(i["saldo"] for i in patrimonio if i["nivel"] == nivel)
    total_patrimonio = total_patrimonio_cuentas + utilidad
    cuadrado = abs(total_activo - (total_pasivo + total_patrimonio)) < Decimal("0.01")

    return {
        "fecha_corte": str(fecha_corte),
        "nivel": nivel,
        "activo": [{"codigo": i["codigo"], "nombre": i["nombre"], "saldo": str(i["saldo"]), "nivel": i["nivel"]} for i in activo],
        "pasivo": [{"codigo": i["codigo"], "nombre": i["nombre"], "saldo": str(i["saldo"]), "nivel": i["nivel"]} for i in pasivo],
        "patrimonio": [{"codigo": i["codigo"], "nombre": i["nombre"], "saldo": str(i["saldo"]), "nivel": i["nivel"]} for i in patrimonio],
        "utilidad_periodo": str(utilidad),
        "totales": {
            "activo":      str(total_activo),
            "pasivo":      str(total_pasivo),
            "patrimonio":  str(total_patrimonio),
            "pasivo_mas_patrimonio": str(total_pasivo + total_patrimonio),
        },
        "cuadrado": cuadrado,
    }


def balance_excel(db: Session, fecha_corte: date, nivel: int = 3):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = balance_general(db, fecha_corte, nivel)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance General"

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num    = '#,##0.00'
    right  = Alignment(horizontal="right")

    ws.merge_cells("A1:C1")
    ws["A1"] = f"Balance General — Corte al {fecha_corte}"
    ws["A1"].font = Font(bold=True, size=12)

    row = 3
    SECCIONES = [
        ("ACTIVO",     data["activo"],     data["totales"]["activo"]),
        ("PASIVO",     data["pasivo"],     data["totales"]["pasivo"]),
        ("PATRIMONIO", data["patrimonio"], None),
    ]

    for seccion_label, items, total in SECCIONES:
        # Encabezado de sección
        c = ws.cell(row=row, column=1, value=seccion_label)
        c.font = Font(bold=True, size=11)
        row += 1

        for item in items:
            indent = (item["nivel"] - 1) * 2  # espacios de sangría
            is_top = item["nivel"] == 1
            ws.cell(row=row, column=1, value=item["codigo"]).border = border
            cn = ws.cell(row=row, column=2, value=(" " * indent) + item["nombre"])
            cn.border = border
            if is_top:
                cn.font = Font(bold=True)
            elif item["nivel"] < nivel:
                cn.font = Font(bold=True, size=9.5)
            c = ws.cell(row=row, column=3, value=float(item["saldo"]))
            c.number_format = num; c.alignment = right; c.border = border
            row += 1

        if seccion_label == "PATRIMONIO":
            # Agregar utilidad del período
            ws.cell(row=row, column=1, value="")
            ws.cell(row=row, column=2, value="Utilidad del período").border = border
            c = ws.cell(row=row, column=3, value=float(data["utilidad_periodo"]))
            c.number_format = num; c.alignment = right; c.border = border
            row += 1
            total = data["totales"]["patrimonio"]

        fill = PatternFill(fill_type="solid", fgColor="EEEEEE")
        etiqueta = f"TOTAL {seccion_label}"
        c1 = ws.cell(row=row, column=1, value=etiqueta)
        c1.font = Font(bold=True); c1.fill = fill; c1.border = border
        ws.cell(row=row, column=2, value="").fill = fill
        c3 = ws.cell(row=row, column=3, value=float(total))
        c3.font = Font(bold=True); c3.number_format = num; c3.alignment = right
        c3.fill = fill; c3.border = border
        row += 2

    # Cuadre final
    fill2 = PatternFill(fill_type="solid", fgColor="D6EAF8")
    ws.cell(row=row, column=2, value="TOTAL PASIVO + PATRIMONIO").font = Font(bold=True, size=11)
    c3 = ws.cell(row=row, column=3, value=float(data["totales"]["pasivo_mas_patrimonio"]))
    c3.font = Font(bold=True, size=11); c3.number_format = num; c3.alignment = right
    for col in [1, 2, 3]:
        ws.cell(row=row, column=col).fill = fill2; ws.cell(row=row, column=col).border = border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"balance_{fecha_corte}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def balanza_excel(db: Session, fecha_desde: date, fecha_hasta: date, nivel: int = 3):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = balanza_comprobacion(db, fecha_desde, fecha_hasta, nivel)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balanza"

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfont  = Font(bold=True, size=10)
    num    = '#,##0.00'
    right  = Alignment(horizontal="right")

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Balanza de Comprobación — {fecha_desde} a {fecha_hasta}"
    ws["A1"].font = Font(bold=True, size=12)

    headers = ["Código", "Nombre", "S.Ini Débito", "S.Ini Crédito",
               "Mov. Débito", "Mov. Crédito", "S.Final Débito", "S.Final Crédito"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hfont; c.border = border
        c.alignment = Alignment(horizontal="right" if col > 2 else "left")

    for ri, f in enumerate(data["filas"], 4):
        vals = [f["codigo"], f["nombre"],
                float(f["si_debito"]), float(f["si_credito"]),
                float(f["periodo_debito"]), float(f["periodo_credito"]),
                float(f["sf_debito"]), float(f["sf_credito"])]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border = border
            if col > 2:
                c.number_format = num; c.alignment = right

    tot_row = len(data["filas"]) + 4
    fill = PatternFill(fill_type="solid", fgColor="EEEEEE")
    t = data["totales"]
    totals = ["TOTALES", "",
              float(t["si_debito"]), float(t["si_credito"]),
              float(t["periodo_debito"]), float(t["periodo_credito"]),
              float(t["sf_debito"]), float(t["sf_credito"])]
    for col, val in enumerate(totals, 1):
        c = ws.cell(row=tot_row, column=col, value=val)
        c.font = Font(bold=True, size=10); c.fill = fill; c.border = border
        if col > 2:
            c.number_format = num; c.alignment = right

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"balanza_{fecha_desde}_{fecha_hasta}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
