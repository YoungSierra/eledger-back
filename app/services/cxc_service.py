import uuid
from datetime import datetime, date, timezone
from decimal import Decimal

from fastapi import HTTPException
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.admin import AdmConfiguracion, AdmConsecutivo, AdmMoneda, AdmTipoDocumento
from app.models.adm import AdmTercero
from app.models.contabilidad import CntAsiento, CntAsientoLinea, CntCuenta, CntPeriodo
from app.models.bancos import BanCuenta
from app.models.cxc import CxcDocumento, CxcRetencion, CxcAplicacion, CxcParametroContable
from app.services.asientos_service import _generar_documento_numero
from app.schemas.auth import UsuarioActual
from app.schemas.cxc import (
    CxcDocumentoCreate, CxcDocumentoUpdate, AnularRequest, AplicarRequest,
    CxcDocumentoResponse, CxcDocumentoListItem, CxcListResponse, RetencionResponse, RetencionCreate,
    CxcResumenItem, CxcResumenResponse,
    ReciboCreate, FacturaPendienteItem, AplicacionPendienteItem,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

TIPO_A_CODIGO = {
    "FACTURA":      "FAC",
    "RECIBO":       "REC",
    "NOTA_CREDITO": "NCC",
    "NOTA_DEBITO":  "NDB",
    "ANTICIPO":     "ANT",
}


def _buscar_periodo(db: Session, fecha: date) -> CntPeriodo:
    p = db.query(CntPeriodo).filter(
        CntPeriodo.fecha_inicio <= fecha,
        CntPeriodo.fecha_cierre >= fecha,
        CntPeriodo.activo == True,
    ).first()
    if not p:
        raise HTTPException(status_code=400, detail=f"No existe período contable para la fecha {fecha}")
    return p


def _moneda_funcional(db: Session) -> AdmMoneda:
    m = db.query(AdmMoneda).filter(AdmMoneda.es_funcional == True, AdmMoneda.activo == True).first()
    if not m:
        raise HTTPException(status_code=400, detail="No hay moneda funcional configurada")
    return m


def _get_config(db: Session, clave: str) -> str:
    cfg = db.query(AdmConfiguracion).filter(AdmConfiguracion.clave == clave).first()
    return cfg.valor if cfg else ""


def _get_cuenta_by_code(db: Session, codigo: str) -> CntCuenta | None:
    if not codigo:
        return None
    return db.query(CntCuenta).filter(CntCuenta.codigo == codigo, CntCuenta.activo == True).first()


def _generar_o_validar_numero(db: Session, tipo: str, numero_manual: str | None) -> str:
    """
    Si numero_manual es None → genera del consecutivo.
    Si viene → valida unicidad y actualiza el consecutivo si el número es mayor al actual.
    """
    codigo = TIPO_A_CODIGO.get(tipo, tipo)
    td = db.query(AdmTipoDocumento).filter(AdmTipoDocumento.codigo == codigo).first()
    if not td:
        raise HTTPException(status_code=400, detail=f"Tipo de documento {codigo} no configurado")

    cons = (
        db.query(AdmConsecutivo)
        .filter(AdmConsecutivo.tipo_documento_id == td.id)
        .with_for_update()
        .first()
    )

    if numero_manual:
        # Validar unicidad
        existe = db.query(CxcDocumento).filter(CxcDocumento.numero == numero_manual).first()
        if existe:
            raise HTTPException(status_code=409, detail=f"El número {numero_manual} ya existe")
        # Actualizar consecutivo si el número parseado es mayor al actual
        if cons:
            sufijo = numero_manual
            if cons.prefijo and sufijo.startswith(cons.prefijo):
                sufijo = sufijo[len(cons.prefijo):]
            try:
                num_parsed = int(sufijo.lstrip("0") or "0")
                if num_parsed > cons.numero_actual:
                    cons.numero_actual = num_parsed
            except ValueError:
                pass
        return numero_manual
    else:
        if not cons:
            raise HTTPException(status_code=400, detail=f"No hay consecutivo configurado para {codigo}")
        siguiente = max(cons.numero_actual + 1, cons.numero_inicio)
        num = str(siguiente).zfill(cons.longitud_minima)
        resultado = f"{cons.prefijo or ''}{num}"
        cons.numero_actual = siguiente
        return resultado


def _calcular_totales(subtotal: Decimal, total_iva: Decimal, total_retenciones: Decimal) -> Decimal:
    return subtotal + total_iva - total_retenciones


def _to_response(doc: CxcDocumento, db: Session) -> CxcDocumentoResponse:
    tercero = db.get(AdmTercero, doc.tercero_id)
    moneda = db.get(AdmMoneda, doc.moneda_id)
    rets = []
    for r in doc.retenciones:
        cuenta = db.get(CntCuenta, r.cuenta_id)
        rets.append(RetencionResponse(
            id=r.id, tipo=r.tipo, concepto=r.concepto,
            base=r.base, porcentaje=r.porcentaje, valor=r.valor,
            cuenta_id=r.cuenta_id,
            cuenta_codigo=cuenta.codigo if cuenta else None,
            cuenta_nombre=cuenta.nombre if cuenta else None,
        ))
    return CxcDocumentoResponse(
        id=doc.id, numero=doc.numero, tipo=doc.tipo,
        fecha=doc.fecha, fecha_vencimiento=doc.fecha_vencimiento,
        periodo_id=doc.periodo_id,
        tercero_id=doc.tercero_id,
        tercero_nit=tercero.nit if tercero else None,
        tercero_nombre=tercero.razon_social if tercero else None,
        moneda_id=doc.moneda_id,
        moneda_codigo=moneda.codigo if moneda else "",
        trm=doc.trm,
        subtotal=doc.subtotal, total_iva=doc.total_iva,
        total_retenciones=doc.total_retenciones,
        total=doc.total, saldo=doc.saldo,
        descripcion=doc.descripcion,
        estado=doc.estado,
        tarifa_iva_id=doc.tarifa_iva_id,
        condicion_pago_id=doc.condicion_pago_id,
        asiento_id=doc.asiento_id,
        asiento_modificado_manual=doc.asiento_modificado_manual,
        documento_origen_id=doc.documento_origen_id,
        ban_cuenta_id=doc.ban_cuenta_id,
        retenciones=rets,
        creado_en=doc.creado_en,
        creado_por=doc.creado_por,
    )


def _to_list_item(doc: CxcDocumento, db: Session, hoy: date) -> CxcDocumentoListItem:
    tercero = db.get(AdmTercero, doc.tercero_id)
    moneda = db.get(AdmMoneda, doc.moneda_id)
    dias = None
    if doc.fecha_vencimiento and doc.estado == "contabilizado" and doc.saldo > 0:
        dias = (doc.fecha_vencimiento - hoy).days
    return CxcDocumentoListItem(
        id=doc.id, numero=doc.numero, tipo=doc.tipo,
        fecha=doc.fecha, fecha_vencimiento=doc.fecha_vencimiento,
        tercero_nit=tercero.nit if tercero else None,
        tercero_nombre=tercero.razon_social if tercero else None,
        moneda_codigo=moneda.codigo if moneda else "",
        total=doc.total, saldo=doc.saldo,
        estado=doc.estado,
        dias_vencimiento=dias,
    )


def _poblar_lineas_asiento_cxc(
    db: Session, asiento_id: uuid.UUID, doc: CxcDocumento,
    cuenta_cxc: CntCuenta, cuenta_ingresos: CntCuenta | None,
    cuenta_iva_p: CntCuenta | None, moneda_func: "AdmMoneda",
) -> None:
    """Genera y persiste las líneas del asiento para un documento CxC."""
    trm = doc.trm or Decimal("1")
    orden = 1

    def add_linea(cuenta_id, debito, credito):
        nonlocal orden
        d_func = (debito * trm).quantize(Decimal("0.0001")) if doc.moneda_id != moneda_func.id else debito
        c_func = (credito * trm).quantize(Decimal("0.0001")) if doc.moneda_id != moneda_func.id else credito
        db.add(CntAsientoLinea(
            id=uuid.uuid4(), asiento_id=asiento_id, orden=orden,
            cuenta_id=cuenta_id,
            debito=debito, credito=credito,
            debito_funcional=d_func, credito_funcional=c_func,
            tercero_id=doc.tercero_id,
        ))
        orden += 1

    if doc.tipo == "FACTURA":
        add_linea(cuenta_cxc.id,      doc.total,    Decimal("0"))
        add_linea(cuenta_ingresos.id, Decimal("0"), doc.subtotal)
        if doc.total_iva > 0 and cuenta_iva_p:
            add_linea(cuenta_iva_p.id, Decimal("0"), doc.total_iva)
        for r in doc.retenciones:
            add_linea(r.cuenta_id, Decimal("0"), r.valor)

    elif doc.tipo == "RECIBO":
        add_linea(cuenta_cxc.id, Decimal("0"), doc.total)
        if doc.ban_cuenta_id:
            ban_cuenta = db.get(BanCuenta, doc.ban_cuenta_id)
            if ban_cuenta and ban_cuenta.cuenta_contable_id:
                add_linea(ban_cuenta.cuenta_contable_id, doc.subtotal, Decimal("0"))
        for r in doc.retenciones:
            add_linea(r.cuenta_id, r.valor, Decimal("0"))

    elif doc.tipo == "NOTA_CREDITO":
        add_linea(cuenta_cxc.id,      Decimal("0"), doc.total)
        add_linea(cuenta_ingresos.id, doc.total,    Decimal("0"))

    elif doc.tipo == "NOTA_DEBITO":
        add_linea(cuenta_cxc.id,      doc.total,    Decimal("0"))
        add_linea(cuenta_ingresos.id, Decimal("0"), doc.total)


def _resolver_parametros_cxc(db: Session, tipo: str):
    """Devuelve (cuenta_cxc, cuenta_ingresos, cuenta_iva_p) o None si faltan las requeridas."""
    params = db.query(CxcParametroContable).first()
    if not params:
        return None
    cuenta_cxc      = db.get(CntCuenta, params.cuenta_clientes_id) if params.cuenta_clientes_id else None
    cuenta_ingresos = db.get(CntCuenta, params.cuenta_ingresos_id) if params.cuenta_ingresos_id else None
    cuenta_iva_p    = db.get(CntCuenta, params.cuenta_iva_id)      if params.cuenta_iva_id      else None
    if not cuenta_cxc:
        return None
    if tipo != "RECIBO" and not cuenta_ingresos:
        return None
    return cuenta_cxc, cuenta_ingresos, cuenta_iva_p


def _generar_asiento_cxc(db: Session, doc: CxcDocumento, actor: UsuarioActual) -> CntAsiento | None:
    """Crea el encabezado del asiento y genera sus líneas."""
    result = _resolver_parametros_cxc(db, doc.tipo)
    if not result:
        return None
    cuenta_cxc, cuenta_ingresos, cuenta_iva_p = result
    moneda_func = _moneda_funcional(db)

    td_codigo = TIPO_A_CODIGO.get(doc.tipo, doc.tipo)
    td = db.query(AdmTipoDocumento).filter(AdmTipoDocumento.codigo == td_codigo).first()

    asiento = CntAsiento(
        id=uuid.uuid4(),
        tipo_documento_id=td.id if td else None,
        documento_numero=doc.numero,
        fecha=doc.fecha,
        periodo_id=doc.periodo_id,
        descripcion=f"{doc.tipo} {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}",
        estado="borrador",
        moneda_id=doc.moneda_id,
        trm=doc.trm if doc.moneda_id != moneda_func.id else None,
        documento_origen_id=doc.id,
        documento_origen_tipo="cxc_documento",
        creado_por=uuid.UUID(actor.id),
    )
    db.add(asiento)
    db.flush()

    _poblar_lineas_asiento_cxc(db, asiento.id, doc, cuenta_cxc, cuenta_ingresos, cuenta_iva_p, moneda_func)
    return asiento


def _get_tercero_nombre(db: Session, tercero_id: uuid.UUID) -> str:
    t = db.get(AdmTercero, tercero_id)
    return t.razon_social if t else ""


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------

def listar(
    db: Session,
    pagina: int = 1,
    por_pagina: int = 50,
    tipo: str | None = None,
    estado: str | None = None,
    tercero_id: uuid.UUID | None = None,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    solo_pendientes: bool = False,
) -> CxcListResponse:
    q = db.query(CxcDocumento).filter(CxcDocumento.activo == True)
    if tipo:       q = q.filter(CxcDocumento.tipo == tipo)
    if estado:     q = q.filter(CxcDocumento.estado == estado)
    if tercero_id: q = q.filter(CxcDocumento.tercero_id == tercero_id)
    if fecha_desde:q = q.filter(CxcDocumento.fecha >= fecha_desde)
    if fecha_hasta:q = q.filter(CxcDocumento.fecha <= fecha_hasta)
    if solo_pendientes:
        q = q.filter(CxcDocumento.saldo > 0, CxcDocumento.estado == "contabilizado")

    total = q.count()
    hoy = date.today()
    rows = q.order_by(CxcDocumento.fecha.desc()).offset((pagina - 1) * por_pagina).limit(por_pagina).all()
    return CxcListResponse(
        items=[_to_list_item(r, db, hoy) for r in rows],
        total=total, pagina=pagina, por_pagina=por_pagina,
    )


def obtener(db: Session, id: uuid.UUID) -> CxcDocumentoResponse:
    doc = db.query(CxcDocumento).filter(CxcDocumento.id == id, CxcDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return _to_response(doc, db)


def crear(db: Session, data: CxcDocumentoCreate, actor: UsuarioActual) -> CxcDocumentoResponse:
    periodo = _buscar_periodo(db, data.fecha)
    moneda_func = _moneda_funcional(db)

    if data.moneda_id != moneda_func.id and not data.trm:
        raise HTTPException(status_code=400, detail="Se requiere TRM para moneda extranjera")

    existe = db.query(CxcDocumento).filter(CxcDocumento.numero == data.numero).first()
    if existe:
        raise HTTPException(status_code=409, detail=f"El número {data.numero} ya existe")

    total = _calcular_totales(data.subtotal, data.total_iva, data.total_retenciones)

    doc = CxcDocumento(
        id=uuid.uuid4(),
        numero=data.numero,
        tipo=data.tipo,
        fecha=data.fecha,
        fecha_vencimiento=data.fecha_vencimiento,
        periodo_id=periodo.id,
        tercero_id=data.tercero_id,
        moneda_id=data.moneda_id,
        trm=data.trm if data.moneda_id != moneda_func.id else None,
        subtotal=data.subtotal,
        total_iva=data.total_iva,
        total_retenciones=data.total_retenciones,
        total=total,
        saldo=total,
        descripcion=data.descripcion,
        tarifa_iva_id=data.tarifa_iva_id,
        condicion_pago_id=data.condicion_pago_id,
        estado="borrador",
        creado_por=uuid.UUID(actor.id),
    )
    db.add(doc)
    db.flush()

    for ret in data.retenciones:
        db.add(CxcRetencion(
            id=uuid.uuid4(), documento_id=doc.id,
            tipo=ret.tipo, concepto=ret.concepto,
            base=ret.base, porcentaje=ret.porcentaje, valor=ret.valor,
            cuenta_id=ret.cuenta_id,
        ))

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def actualizar(db: Session, id: uuid.UUID, data: CxcDocumentoUpdate, actor: UsuarioActual) -> CxcDocumentoResponse:
    doc = db.query(CxcDocumento).filter(CxcDocumento.id == id, CxcDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.estado != "borrador":
        raise HTTPException(status_code=409, detail="Solo se pueden editar documentos en borrador")

    if data.fecha is not None:
        periodo = _buscar_periodo(db, data.fecha)
        doc.fecha = data.fecha
        doc.periodo_id = periodo.id
    if data.fecha_vencimiento is not None: doc.fecha_vencimiento = data.fecha_vencimiento
    if data.tercero_id is not None:        doc.tercero_id = data.tercero_id
    if data.moneda_id is not None:         doc.moneda_id = data.moneda_id
    if data.trm is not None:               doc.trm = data.trm
    if data.descripcion is not None:       doc.descripcion = data.descripcion
    if data.tarifa_iva_id is not None:     doc.tarifa_iva_id = data.tarifa_iva_id
    if data.condicion_pago_id is not None: doc.condicion_pago_id = data.condicion_pago_id

    if any(v is not None for v in [data.subtotal, data.total_iva, data.total_retenciones]):
        subtotal = data.subtotal if data.subtotal is not None else doc.subtotal
        total_iva = data.total_iva if data.total_iva is not None else doc.total_iva
        total_ret = data.total_retenciones if data.total_retenciones is not None else doc.total_retenciones
        total = _calcular_totales(subtotal, total_iva, total_ret)
        doc.subtotal = subtotal
        doc.total_iva = total_iva
        doc.total_retenciones = total_ret
        doc.total = total
        doc.saldo = total

    if data.retenciones is not None:
        db.query(CxcRetencion).filter(CxcRetencion.documento_id == id).delete()
        for ret in data.retenciones:
            db.add(CxcRetencion(
                id=uuid.uuid4(), documento_id=doc.id,
                tipo=ret.tipo, concepto=ret.concepto,
                base=ret.base, porcentaje=ret.porcentaje, valor=ret.valor,
                cuenta_id=ret.cuenta_id,
            ))
        db.flush()

    doc.modificado_por = uuid.UUID(actor.id)
    doc.modificado_en = datetime.now(timezone.utc)
    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def contabilizar(db: Session, id: uuid.UUID, actor: UsuarioActual) -> CxcDocumentoResponse:
    doc = db.query(CxcDocumento).filter(CxcDocumento.id == id, CxcDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.estado != "borrador":
        raise HTTPException(status_code=409, detail="El documento ya está contabilizado o anulado")

    periodo = db.get(CntPeriodo, doc.periodo_id)
    if not periodo or periodo.estado != "abierto":
        raise HTTPException(status_code=400, detail="El período contable no está abierto")

    if doc.total <= 0:
        raise HTTPException(status_code=400, detail="El total del documento debe ser mayor que cero")

    if doc.tipo in ("FACTURA", "NOTA_DEBITO") and not doc.fecha_vencimiento:
        raise HTTPException(status_code=400, detail="La fecha de vencimiento es obligatoria")

    # Para RECIBO: procesar aplicaciones pendientes y actualizar saldos de facturas
    if doc.tipo == "RECIBO":
        apps = db.query(CxcAplicacion).filter(
            CxcAplicacion.documento_credito_id == id,
            CxcAplicacion.estado == "pendiente",
        ).all()
        if not apps:
            raise HTTPException(status_code=400, detail="El recibo no tiene facturas aplicadas")
        for ap in apps:
            fac = db.query(CxcDocumento).filter(
                CxcDocumento.id == ap.documento_debito_id,
                CxcDocumento.activo == True,
            ).with_for_update().first()
            if not fac:
                raise HTTPException(status_code=400, detail="Factura aplicada no encontrada")
            if ap.valor > fac.saldo:
                raise HTTPException(
                    status_code=400,
                    detail=f"El saldo de la factura {fac.numero} cambió ({fac.saldo}). Revisa el recibo."
                )
            fac.saldo -= ap.valor
            ap.estado = "aplicado"
        doc.saldo = Decimal("0")

    # Generar asiento si aún no existe (ej. recibos creados antes de parametrizar cuentas)
    if not doc.asiento_id:
        asiento = _generar_asiento_cxc(db, doc, actor)
        if not asiento:
            raise HTTPException(
                status_code=400,
                detail="No se puede generar el asiento contable. Verifica que las cuentas estén parametrizadas en Administración → Parámetros CxC."
            )
        doc.asiento_id = asiento.id
        db.flush()

    # Publicar el asiento: regenerar líneas desde el recibo para garantizar consistencia
    asiento = db.get(CntAsiento, doc.asiento_id)
    if not asiento:
        raise HTTPException(status_code=400, detail="El asiento contable del documento no fue encontrado")
    if asiento.estado != "borrador":
        raise HTTPException(status_code=409, detail="El asiento ya está publicado o en estado inválido")

    result = _resolver_parametros_cxc(db, doc.tipo)
    if not result:
        raise HTTPException(
            status_code=400,
            detail="No se puede generar el asiento contable. Verifica que las cuentas estén parametrizadas en Administración → Parámetros CxC."
        )
    cuenta_cxc, cuenta_ingresos, cuenta_iva_p = result
    moneda_func = _moneda_funcional(db)

    # Actualizar encabezado y regenerar líneas desde los datos actuales del documento
    asiento.fecha = doc.fecha
    asiento.periodo_id = doc.periodo_id
    asiento.descripcion = f"{doc.tipo} {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}"
    asiento.moneda_id = doc.moneda_id
    asiento.trm = doc.trm if doc.moneda_id != moneda_func.id else None
    db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).delete()
    db.flush()
    _poblar_lineas_asiento_cxc(db, asiento.id, doc, cuenta_cxc, cuenta_ingresos, cuenta_iva_p, moneda_func)
    db.flush()

    # Validar cuadre
    lineas = db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).all()
    total_d = sum(l.debito for l in lineas)
    total_c = sum(l.credito for l in lineas)
    if total_d != total_c:
        raise HTTPException(status_code=400, detail=f"El asiento está descuadrado (débitos {total_d} ≠ créditos {total_c})")

    if asiento.tipo_documento_id and not asiento.documento_numero:
        asiento.documento_numero = _generar_documento_numero(db, asiento.tipo_documento_id)
    asiento.estado = "publicado"
    asiento.modificado_por = uuid.UUID(actor.id)
    asiento.modificado_en = datetime.now(timezone.utc)

    doc.estado = "contabilizado"
    if doc.tipo != "RECIBO":
        doc.saldo = doc.total
    doc.modificado_por = uuid.UUID(actor.id)
    doc.modificado_en = datetime.now(timezone.utc)

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def anular(db: Session, id: uuid.UUID, data: AnularRequest, actor: UsuarioActual) -> CxcDocumentoResponse:
    doc = db.query(CxcDocumento).filter(CxcDocumento.id == id, CxcDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if doc.estado == "anulado":
        raise HTTPException(status_code=409, detail="El documento ya está anulado")

    # Verificar / revertir aplicaciones
    apps_aplicadas = db.query(CxcAplicacion).filter(
        CxcAplicacion.documento_credito_id == id,
        CxcAplicacion.estado == "aplicado",
    ).all()
    if apps_aplicadas:
        raise HTTPException(status_code=409, detail="El recibo ya tiene aplicaciones contabilizadas. Debe anularse mediante contraasiento.")
    # Eliminar aplicaciones pendientes (recibo en borrador)
    db.query(CxcAplicacion).filter(
        CxcAplicacion.documento_credito_id == id,
        CxcAplicacion.estado == "pendiente",
    ).delete()

    doc.estado = "anulado"
    doc.saldo = Decimal("0")
    doc.modificado_por = uuid.UUID(actor.id)
    doc.modificado_en = datetime.now(timezone.utc)
    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def aplicar(db: Session, data: AplicarRequest, actor: UsuarioActual) -> dict:
    credito = db.query(CxcDocumento).filter(CxcDocumento.id == data.documento_credito_id, CxcDocumento.activo == True).first()
    debito = db.query(CxcDocumento).filter(CxcDocumento.id == data.documento_debito_id, CxcDocumento.activo == True).first()

    if not credito or not debito:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if credito.estado != "contabilizado" or debito.estado != "contabilizado":
        raise HTTPException(status_code=409, detail="Ambos documentos deben estar contabilizados")
    if credito.tercero_id != debito.tercero_id:
        raise HTTPException(status_code=400, detail="Los documentos deben pertenecer al mismo tercero")
    if data.valor > credito.saldo:
        raise HTTPException(status_code=400, detail=f"El valor supera el saldo disponible del documento crédito ({credito.saldo})")
    if data.valor > debito.saldo:
        raise HTTPException(status_code=400, detail=f"El valor supera el saldo de la factura ({debito.saldo})")

    db.add(CxcAplicacion(
        id=uuid.uuid4(),
        documento_credito_id=credito.id,
        documento_debito_id=debito.id,
        valor=data.valor,
        fecha=data.fecha,
        creado_por=uuid.UUID(actor.id),
    ))
    credito.saldo -= data.valor
    debito.saldo -= data.valor

    db.commit()
    return {"mensaje": f"Aplicación de {data.valor} registrada correctamente"}


def aplicaciones_pendientes(db: Session, recibo_id: uuid.UUID) -> list[AplicacionPendienteItem]:
    apps = db.query(CxcAplicacion).filter(
        CxcAplicacion.documento_credito_id == recibo_id,
        CxcAplicacion.estado.in_(["pendiente", "aplicado"]),
    ).all()
    result = []
    for ap in apps:
        fac = db.get(CxcDocumento, ap.documento_debito_id)
        if not fac:
            continue
        result.append(AplicacionPendienteItem(
            id=ap.id,
            factura_id=fac.id,
            numero=fac.numero,
            fecha=fac.fecha,
            fecha_vencimiento=fac.fecha_vencimiento,
            total=fac.total,
            saldo_original=fac.saldo + ap.valor,  # saldo real antes de la aplicación pendiente
            valor=ap.valor,
        ))
    return result


def actualizar_recibo(db: Session, recibo_id: uuid.UUID, data: ReciboCreate, actor: UsuarioActual) -> CxcDocumentoResponse:
    doc = db.query(CxcDocumento).filter(CxcDocumento.id == recibo_id, CxcDocumento.activo == True).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Recibo no encontrado")
    if doc.estado != "borrador":
        raise HTTPException(status_code=409, detail="Solo se pueden editar recibos en borrador")

    periodo = _buscar_periodo(db, data.fecha)
    moneda_func = _moneda_funcional(db)

    ban_cuenta = db.get(BanCuenta, data.ban_cuenta_id)
    if not ban_cuenta or not ban_cuenta.activo:
        raise HTTPException(status_code=400, detail="Cuenta bancaria no encontrada")
    if not ban_cuenta.cuenta_contable_id:
        raise HTTPException(status_code=400, detail="La cuenta bancaria no tiene cuenta contable parametrizada")

    total_retenciones = sum(r.valor for r in data.retenciones)
    for ap in data.aplicaciones:
        fac = db.query(CxcDocumento).filter(
            CxcDocumento.id == ap.factura_id,
            CxcDocumento.activo == True,
            CxcDocumento.tipo == "FACTURA",
            CxcDocumento.estado == "contabilizado",
        ).first()
        if not fac:
            raise HTTPException(status_code=400, detail=f"Factura {ap.factura_id} no encontrada")
        if fac.tercero_id != data.tercero_id:
            raise HTTPException(status_code=400, detail=f"La factura {fac.numero} no pertenece al cliente")
        # El saldo disponible de la factura incluye lo que ya tenía pendiente de este recibo
        app_actual = db.query(CxcAplicacion).filter(
            CxcAplicacion.documento_credito_id == recibo_id,
            CxcAplicacion.documento_debito_id == ap.factura_id,
            CxcAplicacion.estado == "pendiente",
        ).first()
        saldo_disp = fac.saldo + (app_actual.valor if app_actual else Decimal("0"))
        if ap.valor > saldo_disp:
            raise HTTPException(status_code=400, detail=f"El valor supera el saldo de la factura {fac.numero}")

    total = data.valor_recibido + total_retenciones

    # Actualizar cabecera
    doc.fecha = data.fecha
    doc.periodo_id = periodo.id
    doc.tercero_id = data.tercero_id
    doc.moneda_id = data.moneda_id
    doc.trm = data.trm if data.moneda_id != moneda_func.id else None
    doc.subtotal = data.valor_recibido
    doc.total_iva = Decimal("0")
    doc.total_retenciones = total_retenciones
    doc.total = total
    doc.saldo = total
    doc.descripcion = data.descripcion
    doc.ban_cuenta_id = data.ban_cuenta_id
    doc.modificado_por = uuid.UUID(actor.id)
    doc.modificado_en = __import__("datetime").datetime.now(__import__("datetime").timezone.utc)

    # Reemplazar retenciones
    db.query(CxcRetencion).filter(CxcRetencion.documento_id == recibo_id).delete()
    for ret in data.retenciones:
        db.add(CxcRetencion(
            id=uuid.uuid4(), documento_id=doc.id,
            tipo=ret.tipo, concepto=ret.concepto,
            base=ret.base, porcentaje=ret.porcentaje, valor=ret.valor,
            cuenta_id=ret.cuenta_id,
        ))

    # Reemplazar aplicaciones pendientes
    db.query(CxcAplicacion).filter(
        CxcAplicacion.documento_credito_id == recibo_id,
        CxcAplicacion.estado == "pendiente",
    ).delete()
    for ap in data.aplicaciones:
        db.add(CxcAplicacion(
            id=uuid.uuid4(),
            documento_credito_id=doc.id,
            documento_debito_id=ap.factura_id,
            valor=ap.valor, fecha=data.fecha,
            estado="pendiente",
            creado_por=uuid.UUID(actor.id),
        ))

    # Regenerar líneas del asiento borrador (mantiene el encabezado para no consumir secuenciales)
    db.flush()
    db.refresh(doc)

    asiento = db.get(CntAsiento, doc.asiento_id) if doc.asiento_id else None
    if asiento and asiento.estado == "borrador":
        result = _resolver_parametros_cxc(db, doc.tipo)
        if result:
            cuenta_cxc, cuenta_ingresos, cuenta_iva_p = result
            moneda_func = _moneda_funcional(db)
            asiento.documento_numero = asiento.documento_numero or doc.numero
            asiento.fecha = doc.fecha
            asiento.periodo_id = doc.periodo_id
            asiento.descripcion = f"{doc.tipo} {doc.numero} — {_get_tercero_nombre(db, doc.tercero_id)}"
            asiento.moneda_id = doc.moneda_id
            asiento.trm = doc.trm if doc.moneda_id != moneda_func.id else None
            db.query(CntAsientoLinea).filter(CntAsientoLinea.asiento_id == asiento.id).delete()
            db.flush()
            _poblar_lineas_asiento_cxc(db, asiento.id, doc, cuenta_cxc, cuenta_ingresos, cuenta_iva_p, moneda_func)
    else:
        nuevo = _generar_asiento_cxc(db, doc, actor)
        if nuevo:
            doc.asiento_id = nuevo.id

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def facturas_pendientes(
    db: Session, tercero_id: uuid.UUID,
    excluir_recibo_id: uuid.UUID | None = None,
) -> list[FacturaPendienteItem]:
    hoy = date.today()
    docs = (
        db.query(CxcDocumento)
        .filter(
            CxcDocumento.activo == True,
            CxcDocumento.tercero_id == tercero_id,
            CxcDocumento.tipo == "FACTURA",
            CxcDocumento.estado == "contabilizado",
            CxcDocumento.saldo > 0,
        )
        .order_by(CxcDocumento.fecha_vencimiento.asc())
        .all()
    )
    result = []
    for d in docs:
        # Descontar aplicaciones pendientes de otros recibos, excluyendo el recibo en edición
        q = db.query(func.coalesce(func.sum(CxcAplicacion.valor), Decimal("0"))).filter(
            CxcAplicacion.documento_debito_id == d.id,
            CxcAplicacion.estado == "pendiente",
        )
        if excluir_recibo_id:
            q = q.filter(CxcAplicacion.documento_credito_id != excluir_recibo_id)
        pendiente = q.scalar()
        saldo_disp = d.saldo - pendiente
        if saldo_disp <= 0:
            continue
        dias = (d.fecha_vencimiento - hoy).days if d.fecha_vencimiento else None
        result.append(FacturaPendienteItem(
            id=d.id, numero=d.numero, fecha=d.fecha,
            fecha_vencimiento=d.fecha_vencimiento,
            total=d.total, aplicado=(d.total - d.saldo) + pendiente, saldo=saldo_disp, dias_vencimiento=dias,
        ))
    return result


def crear_recibo(db: Session, data: ReciboCreate, actor: UsuarioActual) -> CxcDocumentoResponse:
    periodo = _buscar_periodo(db, data.fecha)
    moneda_func = _moneda_funcional(db)

    if data.moneda_id != moneda_func.id and not data.trm:
        raise HTTPException(status_code=400, detail="Se requiere TRM para moneda extranjera")

    ban_cuenta = db.get(BanCuenta, data.ban_cuenta_id)
    if not ban_cuenta or not ban_cuenta.activo:
        raise HTTPException(status_code=400, detail="Cuenta bancaria no encontrada")
    if not ban_cuenta.cuenta_contable_id:
        raise HTTPException(status_code=400, detail="La cuenta bancaria no tiene cuenta contable parametrizada")

    # Validar facturas
    total_retenciones = sum(r.valor for r in data.retenciones)
    for ap in data.aplicaciones:
        fac = db.query(CxcDocumento).filter(
            CxcDocumento.id == ap.factura_id,
            CxcDocumento.activo == True,
            CxcDocumento.tipo == "FACTURA",
            CxcDocumento.estado == "contabilizado",
        ).first()
        if not fac:
            raise HTTPException(status_code=400, detail=f"Factura {ap.factura_id} no encontrada o no contabilizada")
        if fac.tercero_id != data.tercero_id:
            raise HTTPException(status_code=400, detail=f"La factura {fac.numero} no pertenece al cliente seleccionado")
        pendiente_otros = db.query(func.coalesce(func.sum(CxcAplicacion.valor), Decimal("0"))).filter(
            CxcAplicacion.documento_debito_id == ap.factura_id,
            CxcAplicacion.estado == "pendiente",
        ).scalar()
        saldo_disp = fac.saldo - pendiente_otros
        if ap.valor > saldo_disp:
            raise HTTPException(status_code=400, detail=f"El valor a aplicar ({ap.valor}) supera el saldo disponible de la factura {fac.numero} ({saldo_disp})")

    numero = _generar_o_validar_numero(db, "RECIBO", None)
    total = data.valor_recibido + total_retenciones

    doc = CxcDocumento(
        id=uuid.uuid4(),
        numero=numero,
        tipo="RECIBO",
        fecha=data.fecha,
        periodo_id=periodo.id,
        tercero_id=data.tercero_id,
        moneda_id=data.moneda_id,
        trm=data.trm if data.moneda_id != moneda_func.id else None,
        subtotal=data.valor_recibido,
        total_iva=Decimal("0"),
        total_retenciones=total_retenciones,
        total=total,
        saldo=total,
        descripcion=data.descripcion,
        ban_cuenta_id=data.ban_cuenta_id,
        estado="borrador",
        creado_por=uuid.UUID(actor.id),
    )
    db.add(doc)
    db.flush()

    for ret in data.retenciones:
        db.add(CxcRetencion(
            id=uuid.uuid4(), documento_id=doc.id,
            tipo=ret.tipo, concepto=ret.concepto,
            base=ret.base, porcentaje=ret.porcentaje, valor=ret.valor,
            cuenta_id=ret.cuenta_id,
        ))

    for ap in data.aplicaciones:
        db.add(CxcAplicacion(
            id=uuid.uuid4(),
            documento_credito_id=doc.id,
            documento_debito_id=ap.factura_id,
            valor=ap.valor, fecha=data.fecha,
            estado="pendiente",
            creado_por=uuid.UUID(actor.id),
        ))

    db.flush()
    db.refresh(doc)

    # Generar asiento borrador
    asiento = _generar_asiento_cxc(db, doc, actor)
    if asiento:
        doc.asiento_id = asiento.id

    db.commit()
    db.refresh(doc)
    return _to_response(doc, db)


def resumen(db: Session, fecha_corte_str: str | None = None) -> CxcResumenResponse:
    from collections import defaultdict

    hoy = date.fromisoformat(fecha_corte_str) if fecha_corte_str else date.today()

    docs = db.query(CxcDocumento).filter(
        CxcDocumento.activo == True,
        CxcDocumento.estado == "contabilizado",
        CxcDocumento.saldo > 0,
    ).all()

    buckets: dict = defaultdict(lambda: {
        "corriente": Decimal("0"), "dias_1_30": Decimal("0"),
        "dias_31_60": Decimal("0"), "dias_61_90": Decimal("0"), "mas_90": Decimal("0"),
    })

    for doc in docs:
        b = buckets[doc.tercero_id]
        if doc.fecha_vencimiento is None or doc.fecha_vencimiento >= hoy:
            b["corriente"] += doc.saldo
        else:
            dias = (hoy - doc.fecha_vencimiento).days
            if dias <= 30:   b["dias_1_30"]  += doc.saldo
            elif dias <= 60: b["dias_31_60"] += doc.saldo
            elif dias <= 90: b["dias_61_90"] += doc.saldo
            else:            b["mas_90"]     += doc.saldo

    items = []
    tot = {"corriente": Decimal("0"), "dias_1_30": Decimal("0"),
           "dias_31_60": Decimal("0"), "dias_61_90": Decimal("0"), "mas_90": Decimal("0")}

    for tercero_id, b in buckets.items():
        tercero = db.get(AdmTercero, tercero_id)
        total = sum(b.values())
        items.append(CxcResumenItem(
            tercero_id=tercero_id,
            tercero_nit=tercero.nit if tercero else None,
            tercero_nombre=tercero.razon_social if tercero else None,
            corriente=b["corriente"], dias_1_30=b["dias_1_30"],
            dias_31_60=b["dias_31_60"], dias_61_90=b["dias_61_90"],
            mas_90=b["mas_90"], total=total,
        ))
        for k in tot: tot[k] += b[k]

    items.sort(key=lambda x: x.mas_90 + x.dias_61_90 + x.dias_31_60, reverse=True)

    return CxcResumenResponse(
        fecha_corte=hoy, items=items,
        total_corriente=tot["corriente"], total_1_30=tot["dias_1_30"],
        total_31_60=tot["dias_31_60"], total_61_90=tot["dias_61_90"],
        total_mas_90=tot["mas_90"],
        total_general=sum(tot.values()),
    )


def resumen_excel(db: Session, fecha_corte_str: str | None = None):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

    data = resumen(db, fecha_corte_str)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen CxC"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=10)
    title_font  = Font(bold=True, size=12)
    num_fmt = '#,##0'

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Resumen de Cartera — Fecha de corte: {data.fecha_corte}"
    ws["A1"].font = title_font

    # Encabezados
    headers = ["NIT", "Cliente", "Corriente", "1 – 30 días", "31 – 60 días", "61 – 90 días", "+ 90 días", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Filas
    for row_idx, item in enumerate(data.items, 4):
        vals = [
            item.tercero_nit or "",
            item.tercero_nombre or "",
            float(item.corriente),
            float(item.dias_1_30),
            float(item.dias_31_60),
            float(item.dias_61_90),
            float(item.mas_90),
            float(item.total),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            if col >= 3:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")

    # Totales
    total_row = len(data.items) + 4
    totals = [
        "TOTAL", "",
        float(data.total_corriente), float(data.total_1_30),
        float(data.total_31_60), float(data.total_61_90),
        float(data.total_mas_90), float(data.total_general),
    ]
    fill = PatternFill(fill_type="solid", fgColor="EEEEEE")
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True, size=10)
        cell.fill = fill
        cell.border = border
        if col >= 3:
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal="right")

    # Anchos de columna
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"resumen_cxc_{data.fecha_corte}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
