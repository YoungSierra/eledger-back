"""Nómina electrónica — solo transmisión (no liquidación).

_eLedger recibe el detalle de pagos del período y lo transmite a la DIAN vía PTH.
El armado real del XML DIAN y el envío al PTH quedan como stub (igual que la
factura electrónica), pendientes de la integración con el operador.
"""
import io
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

from fastapi import HTTPException
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.admin import AdmConsecutivo, AdmTipoDocumento
from app.models.adm import AdmTercero
from app.models.contabilidad import CntPeriodo
from app.models.nomina import NomPeriodo, NomEmpleado, NomEvento
from app.schemas.auth import UsuarioActual
from app.schemas.nomina import (
    NomPeriodoCreate, NomPeriodoUpdate, AnularNominaRequest,
    NomEmpleadoCreate, NomPeriodoResponse, NomPeriodoListItem, NomListResponse,
    ImportarExcelResponse,
)

CODIGO_NOM = "NOM"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _buscar_periodo(db: Session, fecha: date) -> CntPeriodo:
    p = db.query(CntPeriodo).filter(
        CntPeriodo.fecha_inicio <= fecha,
        CntPeriodo.fecha_cierre >= fecha,
        CntPeriodo.activo == True,
    ).first()
    if not p:
        raise HTTPException(status_code=400, detail=f"No existe período contable para la fecha {fecha}")
    return p


def _generar_numero(db: Session) -> str:
    td = db.query(AdmTipoDocumento).filter(AdmTipoDocumento.codigo == CODIGO_NOM).first()
    if not td:
        raise HTTPException(status_code=400, detail="No existe el tipo de documento NOM (nómina). Verifica el seed de consecutivos.")
    cons = db.query(AdmConsecutivo).filter(AdmConsecutivo.tipo_documento_id == td.id).with_for_update().first()
    if not cons:
        raise HTTPException(status_code=400, detail="No hay consecutivo configurado para NOM.")
    siguiente = max(cons.numero_actual + 1, cons.numero_inicio)
    cons.numero_actual = siguiente
    return f"{cons.prefijo or ''}{str(siguiente).zfill(cons.longitud_minima)}"


def _d(v) -> Decimal:
    if v in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _totales_empleado(e) -> tuple[Decimal, Decimal, Decimal]:
    dev = (_d(e.sueldo) + _d(e.auxilio_transporte) + _d(e.horas_extra)
           + _d(e.bonificaciones) + _d(e.comisiones))
    if getattr(e, "devengados_extra", None):
        dev += sum((_d(v) for v in e.devengados_extra.values()), Decimal("0"))
    ded = (_d(e.salud) + _d(e.pension) + _d(e.fondo_solidaridad) + _d(e.retencion_fuente))
    if getattr(e, "deducciones_extra", None):
        ded += sum((_d(v) for v in e.deducciones_extra.values()), Decimal("0"))
    return dev, ded, dev - ded


# ---------------------------------------------------------------------------
# Serialización
# ---------------------------------------------------------------------------

def _to_response(db: Session, per: NomPeriodo) -> NomPeriodoResponse:
    return NomPeriodoResponse.model_validate(per)


def _to_list_item(per: NomPeriodo) -> NomPeriodoListItem:
    return NomPeriodoListItem(
        id=per.id, numero=per.numero, tipo=per.tipo,
        periodo_pago_inicio=per.periodo_pago_inicio, periodo_pago_fin=per.periodo_pago_fin,
        fecha_generacion=per.fecha_generacion,
        empleados_count=len(per.empleados),
        total_devengado=per.total_devengado, total_deducciones=per.total_deducciones,
        total_neto=per.total_neto, estado=per.estado, dian_estado=per.dian_estado,
        creado_en=per.creado_en,
    )


# ---------------------------------------------------------------------------
# Persistencia de líneas
# ---------------------------------------------------------------------------

def _persistir_empleados(db: Session, per: NomPeriodo, empleados: list[NomEmpleadoCreate]) -> tuple[Decimal, Decimal, Decimal]:
    tot_dev = tot_ded = tot_neto = Decimal("0")
    for i, e in enumerate(empleados, start=1):
        dev, ded, neto = _totales_empleado(e)
        db.add(NomEmpleado(
            id=uuid.uuid4(), periodo_id=per.id, orden=i,
            tipo_documento=e.tipo_documento, numero_documento=e.numero_documento,
            primer_nombre=e.primer_nombre, otros_nombres=e.otros_nombres,
            primer_apellido=e.primer_apellido, segundo_apellido=e.segundo_apellido,
            cargo=e.cargo, salario_basico=_d(e.salario_basico), dias_trabajados=_d(e.dias_trabajados),
            sueldo=_d(e.sueldo), auxilio_transporte=_d(e.auxilio_transporte),
            horas_extra=_d(e.horas_extra), bonificaciones=_d(e.bonificaciones), comisiones=_d(e.comisiones),
            devengados_extra=e.devengados_extra,
            salud=_d(e.salud), pension=_d(e.pension), fondo_solidaridad=_d(e.fondo_solidaridad),
            retencion_fuente=_d(e.retencion_fuente), deducciones_extra=e.deducciones_extra,
            total_devengado=dev, total_deducciones=ded, neto=neto,
        ))
        tot_dev += dev; tot_ded += ded; tot_neto += neto
    return tot_dev, tot_ded, tot_neto


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------

def listar(db: Session, pagina: int = 1, por_pagina: int = 50, estado: str | None = None) -> NomListResponse:
    q = db.query(NomPeriodo).filter(NomPeriodo.activo == True)
    if estado:
        q = q.filter(NomPeriodo.estado == estado)
    total = q.count()
    rows = (q.order_by(NomPeriodo.periodo_pago_fin.desc(), NomPeriodo.creado_en.desc())
            .offset((pagina - 1) * por_pagina).limit(por_pagina).all())
    return NomListResponse(items=[_to_list_item(r) for r in rows], total=total, pagina=pagina, por_pagina=por_pagina)


def obtener(db: Session, id: uuid.UUID) -> NomPeriodoResponse:
    per = db.query(NomPeriodo).filter(NomPeriodo.id == id, NomPeriodo.activo == True).first()
    if not per:
        raise HTTPException(status_code=404, detail="Período de nómina no encontrado")
    return _to_response(db, per)


def crear(db: Session, data: NomPeriodoCreate, actor: UsuarioActual) -> NomPeriodoResponse:
    if data.periodo_pago_fin < data.periodo_pago_inicio:
        raise HTTPException(status_code=400, detail="La fecha fin del período no puede ser anterior a la de inicio")
    periodo = _buscar_periodo(db, data.fecha_generacion)
    numero = _generar_numero(db)
    per = NomPeriodo(
        id=uuid.uuid4(), numero=numero, tipo=data.tipo,
        periodo_pago_inicio=data.periodo_pago_inicio, periodo_pago_fin=data.periodo_pago_fin,
        fecha_generacion=data.fecha_generacion, periodo_id=periodo.id,
        notas=data.notas, estado="borrador", creado_por=uuid.UUID(actor.id),
    )
    db.add(per)
    db.flush()
    dev, ded, neto = _persistir_empleados(db, per, data.empleados)
    per.total_devengado, per.total_deducciones, per.total_neto = dev, ded, neto
    db.commit()
    db.refresh(per)
    return _to_response(db, per)


def actualizar(db: Session, id: uuid.UUID, data: NomPeriodoUpdate, actor: UsuarioActual) -> NomPeriodoResponse:
    per = db.query(NomPeriodo).filter(NomPeriodo.id == id, NomPeriodo.activo == True).first()
    if not per:
        raise HTTPException(status_code=404, detail="Período de nómina no encontrado")
    if per.estado != "borrador":
        raise HTTPException(status_code=409, detail="Solo se pueden editar períodos en borrador")

    if data.tipo is not None:                 per.tipo = data.tipo
    if data.periodo_pago_inicio is not None:  per.periodo_pago_inicio = data.periodo_pago_inicio
    if data.periodo_pago_fin is not None:     per.periodo_pago_fin = data.periodo_pago_fin
    if data.fecha_generacion is not None:
        per.fecha_generacion = data.fecha_generacion
        per.periodo_id = _buscar_periodo(db, data.fecha_generacion).id
    if data.notas is not None:                per.notas = data.notas
    if per.periodo_pago_fin < per.periodo_pago_inicio:
        raise HTTPException(status_code=400, detail="La fecha fin del período no puede ser anterior a la de inicio")

    if data.empleados is not None:
        db.query(NomEmpleado).filter(NomEmpleado.periodo_id == per.id).delete()
        db.flush()
        dev, ded, neto = _persistir_empleados(db, per, data.empleados)
        per.total_devengado, per.total_deducciones, per.total_neto = dev, ded, neto

    per.modificado_por = uuid.UUID(actor.id)
    per.modificado_en = datetime.now(timezone.utc)
    db.commit()
    db.refresh(per)
    return _to_response(db, per)


def eliminar(db: Session, id: uuid.UUID, actor: UsuarioActual) -> None:
    per = db.query(NomPeriodo).filter(NomPeriodo.id == id, NomPeriodo.activo == True).first()
    if not per:
        raise HTTPException(status_code=404, detail="Período de nómina no encontrado")
    if per.estado not in ("borrador", "generado"):
        raise HTTPException(status_code=409, detail="Solo se pueden descartar períodos en borrador o generado (aún no enviados)")
    per.activo = False
    per.estado = "anulado"
    per.modificado_por = uuid.UUID(actor.id)
    per.modificado_en = datetime.now(timezone.utc)
    db.commit()


# ---------------------------------------------------------------------------
# Generar / enviar (STUB — pendiente integración PTH, igual que facturas)
# ---------------------------------------------------------------------------

def _validar_empleados_terceros(db: Session, per: NomPeriodo) -> None:
    """Exige que la identificación de cada empleado exista como tercero tipo EMPLEADO.
    Se valida al generar (para transmitir), NO al guardar borrador."""
    docs = [e.numero_documento.strip() for e in per.empleados if e.numero_documento]
    if not docs:
        return
    existentes = {
        t.nit for t in db.query(AdmTercero).filter(
            AdmTercero.nit.in_(docs), AdmTercero.tipo_tercero == "EMPLEADO", AdmTercero.activo == True,
        ).all()
    }
    faltantes = []
    for e in per.empleados:
        doc = (e.numero_documento or "").strip()
        if doc and doc not in existentes and doc not in faltantes:
            faltantes.append(f"{doc} ({e.primer_nombre} {e.primer_apellido})")
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail="Estas identificaciones no existen como empleado (tercero tipo EMPLEADO): "
                   + "; ".join(faltantes) + ". Créalos en Terceros antes de generar la nómina.",
        )


def generar(db: Session, id: uuid.UUID, actor: UsuarioActual) -> NomPeriodoResponse:
    """Arma el XML DIAN (stub) y lo deja listo para transmitir."""
    per = db.query(NomPeriodo).filter(NomPeriodo.id == id, NomPeriodo.activo == True).first()
    if not per:
        raise HTTPException(status_code=404, detail="Período de nómina no encontrado")
    if per.estado not in ("borrador", "generado"):
        raise HTTPException(status_code=409, detail="El período ya fue enviado o anulado")
    if not per.empleados:
        raise HTTPException(status_code=400, detail="El período no tiene empleados")
    _validar_empleados_terceros(db, per)

    xml = _armar_xml_stub(per)
    key = f"nomina/{per.numero}.xml"
    try:
        from app.core.almacenamiento import subir
        subir(key, xml.encode("utf-8"), "application/xml")
        per.xml_key = key
    except Exception:
        # Sin almacenamiento (dev): no bloquea; el XML se puede regenerar.
        per.xml_key = None

    per.estado = "generado"
    per.dian_estado = None
    db.add(NomEvento(id=uuid.uuid4(), periodo_id=per.id, tipo="RESPUESTA", estado="generado",
                     mensaje="XML de nómina generado (stub, pendiente envío a PTH).",
                     creado_por=uuid.UUID(actor.id)))
    per.modificado_por = uuid.UUID(actor.id)
    per.modificado_en = datetime.now(timezone.utc)
    db.commit()
    db.refresh(per)
    return _to_response(db, per)


def enviar(db: Session, id: uuid.UUID, actor: UsuarioActual) -> NomPeriodoResponse:
    """STUB de transmisión al PTH. Cuando exista la integración real, aquí se
    enviará el XML y se procesará la respuesta (CUNE / estado DIAN)."""
    per = db.query(NomPeriodo).filter(NomPeriodo.id == id, NomPeriodo.activo == True).first()
    if not per:
        raise HTTPException(status_code=404, detail="Período de nómina no encontrado")
    if per.estado not in ("generado", "rechazado"):
        raise HTTPException(status_code=409, detail="Primero genera el documento (o solo se reintenta un rechazado)")

    per.estado = "enviado"
    per.dian_estado = "pendiente"
    per.dian_mensaje = "Transmisión pendiente: la integración con el PTH aún no está habilitada."
    db.add(NomEvento(id=uuid.uuid4(), periodo_id=per.id, tipo="ENVIO", estado="pendiente",
                     mensaje="Envío registrado (stub). Falta integración real con el PTH.",
                     creado_por=uuid.UUID(actor.id)))
    per.modificado_por = uuid.UUID(actor.id)
    per.modificado_en = datetime.now(timezone.utc)
    db.commit()
    db.refresh(per)
    return _to_response(db, per)


def anular(db: Session, id: uuid.UUID, data: AnularNominaRequest, actor: UsuarioActual) -> NomPeriodoResponse:
    per = db.query(NomPeriodo).filter(NomPeriodo.id == id, NomPeriodo.activo == True).first()
    if not per:
        raise HTTPException(status_code=404, detail="Período de nómina no encontrado")
    if per.estado == "anulado":
        raise HTTPException(status_code=409, detail="El período ya está anulado")
    per.estado = "anulado"
    per.activo = False
    db.add(NomEvento(id=uuid.uuid4(), periodo_id=per.id, tipo="RESPUESTA", estado="anulado",
                     mensaje=f"Anulado: {data.motivo}", creado_por=uuid.UUID(actor.id)))
    per.modificado_por = uuid.UUID(actor.id)
    per.modificado_en = datetime.now(timezone.utc)
    db.commit()
    db.refresh(per)
    return _to_response(db, per)


def _armar_xml_stub(per: NomPeriodo) -> str:
    """XML mínimo de referencia (NO es el UBL DIAN definitivo). Placeholder hasta
    integrar el esquema oficial de Nómina Electrónica v1.0."""
    def esc(s):
        return (str(s or "")).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    filas = []
    for e in per.empleados:
        filas.append(
            f'    <Trabajador numeroDocumento="{esc(e.numero_documento)}" '
            f'nombre="{esc(e.primer_nombre)} {esc(e.primer_apellido)}" '
            f'devengado="{e.total_devengado}" deducciones="{e.total_deducciones}" neto="{e.neto}"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        f'<NominaIndividual numero="{esc(per.numero)}" tipo="{esc(per.tipo)}" '
        f'periodoInicio="{per.periodo_pago_inicio}" periodoFin="{per.periodo_pago_fin}" '
        f'totalDevengado="{per.total_devengado}" totalDeducciones="{per.total_deducciones}" '
        f'comprobanteTotal="{per.total_neto}">\n'
        '  <Trabajadores>\n' + "\n".join(filas) + '\n  </Trabajadores>\n'
        '</NominaIndividual>\n'
    )


# ---------------------------------------------------------------------------
# Excel: plantilla + importación
# ---------------------------------------------------------------------------

_COLUMNAS = [
    ("tipo_documento", "Tipo doc"),
    ("numero_documento", "N° documento"),
    ("primer_nombre", "Primer nombre"),
    ("otros_nombres", "Otros nombres"),
    ("primer_apellido", "Primer apellido"),
    ("segundo_apellido", "Segundo apellido"),
    ("cargo", "Cargo"),
    ("salario_basico", "Salario básico"),
    ("dias_trabajados", "Días trabajados"),
    ("sueldo", "Sueldo"),
    ("auxilio_transporte", "Aux. transporte"),
    ("horas_extra", "Horas extra"),
    ("bonificaciones", "Bonificaciones"),
    ("comisiones", "Comisiones"),
    ("salud", "Salud"),
    ("pension", "Pensión"),
    ("fondo_solidaridad", "Fondo solidaridad"),
    ("retencion_fuente", "Retención fuente"),
]
_NUMERICAS = {"salario_basico", "dias_trabajados", "sueldo", "auxilio_transporte", "horas_extra",
              "bonificaciones", "comisiones", "salud", "pension", "fondo_solidaridad", "retencion_fuente"}


def generar_plantilla_excel() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"
    ws.append([titulo for _, titulo in _COLUMNAS])
    # Fila de ejemplo
    ws.append(["CC", "1000000", "Juan", "", "Pérez", "", "Auxiliar",
               1300000, 30, 1300000, 162000, 0, 0, 0, 52000, 52000, 0, 0])
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_excel(data: bytes) -> ImportarExcelResponse:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo. Debe ser un Excel (.xlsx).")
    ws = wb.active
    campos = [c for c, _ in _COLUMNAS]
    empleados: list[NomEmpleadoCreate] = []
    avisos: list[str] = []
    filas = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v in (None, "") for v in row):
            continue
        filas += 1
        d = {campos[j]: (row[j] if j < len(row) else None) for j in range(len(campos))}
        if not d.get("numero_documento") or not d.get("primer_nombre") or not d.get("primer_apellido"):
            avisos.append(f"Fila {i}: omitida (faltan documento, primer nombre o primer apellido).")
            continue
        payload = {
            "tipo_documento": str(d.get("tipo_documento") or "CC").strip(),
            "numero_documento": str(d.get("numero_documento")).strip(),
            "primer_nombre": str(d.get("primer_nombre")).strip(),
            "otros_nombres": (str(d["otros_nombres"]).strip() if d.get("otros_nombres") else None),
            "primer_apellido": str(d.get("primer_apellido")).strip(),
            "segundo_apellido": (str(d["segundo_apellido"]).strip() if d.get("segundo_apellido") else None),
            "cargo": (str(d["cargo"]).strip() if d.get("cargo") else None),
        }
        for c in _NUMERICAS:
            payload[c] = _d(d.get(c))
        empleados.append(NomEmpleadoCreate(**payload))
    if not empleados:
        avisos.append("No se importó ningún empleado válido.")
    return ImportarExcelResponse(empleados=empleados, filas_leidas=filas, avisos=avisos)
